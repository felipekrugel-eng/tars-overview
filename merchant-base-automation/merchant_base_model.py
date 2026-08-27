#!/usr/bin/env python3
"""
merchant_base_model.py — add Felipe's own layer back on top of a freshly generated
US Merchant Base workbook.

    python3 merchant_base_model.py --workbook US_Merchant_Base_FULL.xlsx \
        [--previous <last published workbook, for manual-column merge>]

build_full_base.py produces seven tabs from the Snowflake export. Everything else in
the workbook is Felipe's, and this script reproduces it deterministically:

    Margin Assumptions     hand-entered inputs + sourcing notes, and the aggregate
                           outputs. Inputs are written ONCE from ASSUMPTIONS below;
                           on a refresh the previous workbook's input cells win, so
                           editing a price in Excel is not undone by the next run.
    Model columns          Revenue / Margin x Tier / breakeven / working columns,
                           appended to Full Base.
    ACTION                 hand-typed follow-ups, merged forward on MERCHANT_ID.
    Snapshot tabs          Prime Base, ACTIVE L12M, Transacting Merchants,
                           Paying Total, EMAIL BASE - rule-driven, so they stop
                           being frozen pastes.

THREE DESIGN DECISIONS, each fixing something in the hand-built file:

1. FORMULAS ONLY WHERE GTV > 0.
   Every model formula is wrapped in =IF($<GTV>row>0, ..., ""). In the hand-built
   workbook they were filled across all 131,248 rows, but only 14,676 (11.2%) have
   POS_GTV_12M_USD > 0 - so 116,572 rows ran 24 formulas each to return an empty
   string. 2,797,728 cells computing nothing, and the reason Full Base was a 496 MB
   XML part inside a 151 MB file. Output is identical; the file is ~10x smaller.

2. COLUMNS ARE RESOLVED BY NAME, NEVER BY LETTER.
   The hand-built formulas hardcoded $AB (GTV) and $AF (receipts). Those letters were
   only correct because 'GTV MONTH AVERAGE' had been inserted by hand at AD; the
   generator writes straight from the CSV header order and would have shifted
   everything after it, silently re-pointing the whole model. Here every reference is
   looked up from the header row at build time, so any column-order change is safe.
   'GTV MONTH AVERAGE' is now a generated column like any other.

3. RANGES ARE BOUNDED AT MAXROW, NOT AT THE CURRENT ROW COUNT.
   Margin Assumptions bounded every output at row 131250 - exactly the row count on
   the day it was written. The first refresh that added a merchant would have dropped
   it from every total, silently. Same failure as the 16091 ceiling in the margin
   workbook.
"""

import argparse
import pathlib
import sys
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as gcl

MAXROW = 300_000
DATA_SHEET = "Full Base"
HDR_ROW = 1
DATA_START = 2

NAVY = "0B2B3C"
BLUE = "1F6FB2"
INPUT_FILL = "DCE9F5"          # blue = an input you may edit
HDR_F = Font(name="Aptos Narrow", size=10, bold=True, color="FFFFFF")
BODY_F = Font(name="Aptos Narrow", size=10)
TITLE_F = Font(name="Aptos Narrow", size=12, bold=True, color=NAVY)
NOTE_F = Font(name="Aptos Narrow", size=9, italic=True)
INPUT_F = Font(name="Aptos Narrow", size=10, bold=True, color=BLUE)

MONEY = '$#,##0.00;($#,##0.00);-'
PCT4 = '0.0000%;-0.0000%;-'
INT = '#,##0;(#,##0);-'

# ---------------------------------------------------------------------------
# Margin Assumptions. (row, col, value, kind) - kind drives formatting and, for
# 'input', whether the previous workbook's value is carried forward.
# Transcribed from the hand-built workbook, 21 Aug 2026.
# ---------------------------------------------------------------------------
ASSUMPTIONS = [
    (1, 1, "Payments Margin — Assumptions", "title"),
    (3, 1, "Volume inputs", "head"),
    (4, 1, "Card share of GTV", "label"),          (4, 2, 0.80, "input_pct"),
    (5, 1, "Months per year", "label"),            (5, 2, 12, "input_int"),
    (7, 1, "Merchant pricing — gross, % of card value (no fixed fee)", "head"),
    (8, 1, "Price 1", "label"),                    (8, 2, 0.0199, "input_pct"),
    (9, 1, "Price 2", "label"),                    (9, 2, 0.0229, "input_pct"),
    (10, 1, "Price 3", "label"),                   (10, 2, 0.0249, "input_pct"),
    (12, 1, "Network cost — interchange + card scheme + Amex (invariant to tier)", "head"),
    (13, 1, "Network cost, % of card volume", "label"), (13, 2, 0.0212137, "input_pct"),
    (15, 1, "Stripe markup by Loyverse-wide monthly card volume tier", "head"),
    (16, 2, "Tier 1", "colhead"), (16, 3, "Tier 2", "colhead"), (16, 4, "Tier 3", "colhead"),
    (17, 1, "% of card value", "label"),
    (17, 2, 0.0010, "input_pct"), (17, 3, 0.0007, "input_pct"), (17, 4, 0.0005, "input_pct"),
    (18, 1, "Fixed per authorisation (USD)", "label"),
    (18, 2, 0.05, "input_money"), (18, 3, 0.03, "input_money"), (18, 4, 0.025, "input_money"),
    (19, 1, "Loyverse monthly volume band", "label"),
    (19, 2, "0 – $50m", "note"), (19, 3, "> $50m – $100m", "note"), (19, 4, "> $100m", "note"),

    (21, 1, "Notes & sources", "head"),
    (22, 1, "Card share of GTV is a planning assumption supplied by the user (80%). "
            "Applied to BOTH card volume and card transaction count.", "note"),
    (23, 1, "Network cost %: blended actual from "
            "Loyverse_Payments_Transaction_Margin_vs_Competitors, 'Transaction Detail' "
            "- interchange + card scheme + Amex discount over total volume.", "note"),
    (24, 1, "Stripe markup: Stripe Pricing Agreement (Loyverse Commerce Ltd / Five "
            "Galaxies Commerce Ltd, signed 15 Dec 2025), IC+ Credit and Debit Cards "
            "(Tiered), Card-Present fee table.", "note"),
    (25, 1, "TIERS ARE NOT A MERCHANT ATTRIBUTE. T1/T2/T3 are Loyverse's own aggregate "
            "monthly card volume tiers with Stripe. Every merchant is priced at "
            "whichever tier Loyverse is in that month.", "note"),
    (26, 1, "Tiers are charged on a progressive basis within a month, so a blended rate "
            "applies at the boundary. Using a single tier is a simplification.", "note"),
    (27, 1, "NOT MODELLED (per source workbook): chargeback/dispute provision, "
            "terminal hardware, refunds, FX.", "note"),
    (28, 1, "Data caveats on Full Base: some POS_GTV_12M_USD values are capped "
            "(understating volume); see GTV_DATA_FLAG and GTV_CONFIDENCE.", "note"),

    (30, 1, "Average-ticket sanity band (flags implausible rows)", "head"),
    (31, 1, "Lower bound — below this, loss-making at every price and tier", "label"),
    (31, 2, 8, "input_money"),
    (32, 1, "Upper bound — above this, suspect mislabelled currency", "label"),
    (32, 2, 60, "input_money"),
]

# Aggregate outputs. {col} placeholders are resolved to the model column letters.
OUTPUTS = [
    (34, 1, "Model output — monthly, all merchants with GTV > 0", "head"),
    (35, 1, "Merchants with GTV > 0", "label"),
    (35, 2, '=COUNTIF({GTV_R},">0")', "int"),
    (36, 1, "Monthly card volume (USD)", "label"), (36, 2, "=SUM({CI_R})", "money"),
    (37, 1, "Monthly card transactions", "label"), (37, 2, "=SUM({CJ_R})", "int"),
    (38, 1, "Blended average card ticket (USD)", "label"),
    (38, 2, '=IFERROR(B36/B37,"")', "money"),

    (40, 1, "Monthly revenue (USD)", "head"),
    (41, 1, "at Price 1", "label"), (41, 2, "=SUM({BQ_R})", "money"),
    (42, 1, "at Price 2", "label"), (42, 2, "=SUM({BR_R})", "money"),
    (43, 1, "at Price 3", "label"), (43, 2, "=SUM({BS_R})", "money"),

    (45, 1, "Monthly cost (USD) by Loyverse tier scenario", "head"),
    (46, 2, "Tier 1", "colhead"), (46, 3, "Tier 2", "colhead"), (46, 4, "Tier 3", "colhead"),
    (47, 1, "Total cost", "label"),
    (47, 2, "=SUM({CL_R})", "money"), (47, 3, "=SUM({CM_R})", "money"),
    (47, 4, "=SUM({CN_R})", "money"),
    (48, 1, "Blended breakeven price", "label"),
    (48, 2, '=IFERROR(B47/$B$36,"")', "pct"), (48, 3, '=IFERROR(C47/$B$36,"")', "pct"),
    (48, 4, '=IFERROR(D47/$B$36,"")', "pct"),

    (50, 1, "Monthly margin (USD) — price x tier", "head"),
    (51, 2, "Tier 1", "colhead"), (51, 3, "Tier 2", "colhead"), (51, 4, "Tier 3", "colhead"),
    (52, 1, "=B8", "pctlabel"),
    (52, 2, "=SUM({BV_R})", "money"), (52, 3, "=SUM({BW_R})", "money"),
    (52, 4, "=SUM({BX_R})", "money"),
    (53, 1, "=B9", "pctlabel"),
    (53, 2, "=SUM({BY_R})", "money"), (53, 3, "=SUM({BZ_R})", "money"),
    (53, 4, "=SUM({CA_R})", "money"),
    (54, 1, "=B10", "pctlabel"),
    (54, 2, "=SUM({CB_R})", "money"), (54, 3, "=SUM({CC_R})", "money"),
    (54, 4, "=SUM({CD_R})", "money"),

    (56, 1, "Model output — monthly, EXCLUDING rows flagged by the ticket sanity band",
     "head"),
    (57, 1, "Merchants included", "label"), (57, 2, "=SUM({CP_R})", "int"),
    (58, 1, "Monthly card volume (USD)", "label"),
    (58, 2, "=SUMIF({CP_R},1,{CI_R})", "money"),
    (59, 1, "Monthly card transactions", "label"),
    (59, 2, "=SUMIF({CP_R},1,{CJ_R})", "int"),
    (60, 1, "Blended average card ticket (USD)", "label"),
    (60, 2, '=IFERROR(B58/B59,"")', "money"),

    (62, 1, "Monthly revenue (USD)", "head"),
    (63, 1, "at Price 1", "label"), (63, 2, "=SUMIF({CP_R},1,{BQ_R})", "money"),
    (64, 1, "at Price 2", "label"), (64, 2, "=SUMIF({CP_R},1,{BR_R})", "money"),
    (65, 1, "at Price 3", "label"), (65, 2, "=SUMIF({CP_R},1,{BS_R})", "money"),

    (67, 1, "Monthly cost (USD)", "head"),
    (67, 2, "Tier 1", "colhead"), (67, 3, "Tier 2", "colhead"), (67, 4, "Tier 3", "colhead"),
    (68, 1, "Total cost", "label"),
    (68, 2, "=SUMIF({CP_R},1,{CL_R})", "money"),
    (68, 3, "=SUMIF({CP_R},1,{CM_R})", "money"),
    (68, 4, "=SUMIF({CP_R},1,{CN_R})", "money"),
    (69, 1, "Blended breakeven price", "label"),
    (69, 2, '=IFERROR(B68/$B$58,"")', "pct"), (69, 3, '=IFERROR(C68/$B$58,"")', "pct"),
    (69, 4, '=IFERROR(D68/$B$58,"")', "pct"),

    (71, 1, "Monthly margin (USD)", "head"),
    (71, 2, "Tier 1", "colhead"), (71, 3, "Tier 2", "colhead"), (71, 4, "Tier 3", "colhead"),
    (72, 1, "=B8", "pctlabel"),
    (72, 2, "=SUMIF({CP_R},1,{BV_R})", "money"),
    (72, 3, "=SUMIF({CP_R},1,{BW_R})", "money"),
    (72, 4, "=SUMIF({CP_R},1,{BX_R})", "money"),
    (73, 1, "=B9", "pctlabel"),
    (73, 2, "=SUMIF({CP_R},1,{BY_R})", "money"),
    (73, 3, "=SUMIF({CP_R},1,{BZ_R})", "money"),
    (73, 4, "=SUMIF({CP_R},1,{CA_R})", "money"),
    (74, 1, "=B10", "pctlabel"),
    (74, 2, "=SUMIF({CP_R},1,{CB_R})", "money"),
    (74, 3, "=SUMIF({CP_R},1,{CC_R})", "money"),
    (74, 4, "=SUMIF({CP_R},1,{CD_R})", "money"),
]

INPUT_CELLS = [f"{gcl(c)}{r}" for r, c, _v, k in ASSUMPTIONS if k.startswith("input")]

A = "'Margin Assumptions'!"

# Model columns appended to Full Base. key -> (header, number format, formula).
# {G} = GTV column letter, {R} = receipts column letter, {row} = the row number.
# Headers embed the old banner text ("Revenue", "Margin", "Working columns") on a
# second line, because the generator writes a single header row and inserting a
# banner row above 131k rows of data would shift every reference in the workbook.
MODEL = OrderedDict([
    ("BQ", ("Revenue\nPrice 1", MONEY, f'=IF(${{G}}{{row}}>0,$CI{{row}}*{A}$B$8,"")')),
    ("BR", ("Revenue\nPrice 2", MONEY, f'=IF(${{G}}{{row}}>0,$CI{{row}}*{A}$B$9,"")')),
    ("BS", ("Revenue\nPrice 3", MONEY, f'=IF(${{G}}{{row}}>0,$CI{{row}}*{A}$B$10,"")')),
    ("BT", ("Minimum Breakeven Price", PCT4, '=IFERROR($CL{row}/$CI{row},"")')),
    ("BU", (None, None, None)),
    ("BV", ("Margin\nPrice 1 (T1)", MONEY, '=IF(${G}{row}>0,$BQ{row}-$CL{row},"")')),
    ("BW", ("Margin\nPrice 1 (T2)", MONEY, '=IF(${G}{row}>0,$BQ{row}-$CM{row},"")')),
    ("BX", ("Margin\nPrice 1 (T3)", MONEY, '=IF(${G}{row}>0,$BQ{row}-$CN{row},"")')),
    ("BY", ("Margin\nPrice 2 (T1)", MONEY, '=IF(${G}{row}>0,$BR{row}-$CL{row},"")')),
    ("BZ", ("Margin\nPrice 2 (T2)", MONEY, '=IF(${G}{row}>0,$BR{row}-$CM{row},"")')),
    ("CA", ("Margin\nPrice 2 (T3)", MONEY, '=IF(${G}{row}>0,$BR{row}-$CN{row},"")')),
    ("CB", ("Margin\nPrice 3 (T1)", MONEY, '=IF(${G}{row}>0,$BS{row}-$CL{row},"")')),
    ("CC", ("Margin\nPrice 3 (T2)", MONEY, '=IF(${G}{row}>0,$BS{row}-$CM{row},"")')),
    ("CD", ("Margin\nPrice 3 (T3)", MONEY, '=IF(${G}{row}>0,$BS{row}-$CN{row},"")')),
    ("CE", ("Minimum Breakeven Price (T1)", PCT4, '=IFERROR($CL{row}/$CI{row},"")')),
    ("CF", ("Minimum Breakeven Price (T2)", PCT4, '=IFERROR($CM{row}/$CI{row},"")')),
    ("CG", ("Minimum Breakeven Price (T3)", PCT4, '=IFERROR($CN{row}/$CI{row},"")')),
    ("CH", (None, None, None)),
    ("CI", ("Working columns\nMonthly card volume (USD)", MONEY,
            f'=IF(${{G}}{{row}}>0,${{G}}{{row}}/{A}$B$5*{A}$B$4,"")')),
    ("CJ", ("Working columns\nMonthly card transactions", '0.00',
            f'=IF(${{G}}{{row}}>0,${{R}}{{row}}/{A}$B$5*{A}$B$4,"")')),
    ("CK", ("Working columns\nAvg card ticket (USD)", MONEY,
            '=IFERROR($CI{row}/$CJ{row},"")')),
    ("CL", ("Working columns\nMonthly cost — T1 (USD)", MONEY,
            f'=IF(${{G}}{{row}}>0,$CI{{row}}*({A}$B$13+{A}B$17)+$CJ{{row}}*{A}B$18,"")')),
    ("CM", ("Working columns\nMonthly cost — T2 (USD)", MONEY,
            f'=IF(${{G}}{{row}}>0,$CI{{row}}*({A}$B$13+{A}C$17)+$CJ{{row}}*{A}C$18,"")')),
    ("CN", ("Working columns\nMonthly cost — T3 (USD)", MONEY,
            f'=IF(${{G}}{{row}}>0,$CI{{row}}*({A}$B$13+{A}D$17)+$CJ{{row}}*{A}D$18,"")')),
    ("CO", ("Working columns\nTicket sanity flag", None,
            f'=IF(${{G}}{{row}}=0,"",IF($CK{{row}}<{A}$B$31,'
            f'"Ticket below band — loss-making at every price",'
            f'IF($CK{{row}}>{A}$B$32,'
            f'"Ticket above band — check for mislabelled currency","")))')),
    ("CP", ("Working columns\nClean row (1/0)", INT,
            '=IF(${G}{row}>0,IF($CO{row}="",1,0),0)')),
])

# ---------------------------------------------------------------------------
# Snapshot tabs. Each is a RULE over Full Base, so they refresh with the data
# instead of being frozen pastes.
#
# Prime Base was reverse-engineered EXACTLY against the 21 Aug workbook: the rule
# below reproduces its 939 rows with zero extras and zero omissions.
#
# The other three are best-fit. Each rule is a strict SUPERSET of the hand-built tab
# (0 rows missing, some extra), and no column in Full Base separates the extras - so
# those tabs were pruned by hand or exported with a filter not recorded anywhere.
# Row counts at 21 Aug: rule vs tab -> Transacting 52 v 50, ACTIVE 1,900 v 1,716,
# Paying Total 2,540 v 1,954. Automating them WIDENS the tabs. Confirm that is right.
# ---------------------------------------------------------------------------
def _yes(v):
    return str(v).strip().lower() == "yes"


SNAPSHOTS = OrderedDict([
    ("Prime Base", dict(
        note="Active on POS, paying, targetable, contactable, and GTV data we trust. "
             "Reproduces the 21 Aug hand-built tab exactly (939 rows).",
        rule=lambda r: (_yes(r.get("ACTIVE_ON_POS")) and _yes(r.get("IS_PAYING"))
                        and _yes(r.get("IS_TARGETABLE")) and _yes(r.get("IS_CONTACTABLE"))
                        and str(r.get("GTV_CONFIDENCE")) in
                        ("2 - Plausible", "No recent till volume")),
        sort=lambda r: -(_num(r.get("POS_GTV_12M_USD"))),
    )),
    ("ACTIVE L12M_GTV 2500-28000", dict(
        note="Monthly average uncapped GTV between $2,500 and $28,000, contactable. "
             "BEST FIT: 1,900 rows vs 1,716 in the hand-built tab.",
        rule=lambda r: (_num(r.get("POS_GTV_12M_UNCAPPED_USD")) / 12 > 2500
                        and _num(r.get("POS_GTV_12M_UNCAPPED_USD")) / 12 < 28000
                        and _yes(r.get("IS_CONTACTABLE"))),
        sort=lambda r: -(_num(r.get("POS_GTV_12M_UNCAPPED_USD"))),
    )),
    ("Transacting Merchants", dict(
        note="FUNNEL_STAGE = Transacting. BEST FIT: 52 rows vs 50 in the hand-built tab.",
        rule=lambda r: str(r.get("FUNNEL_STAGE")).strip() == "Transacting",
        sort=lambda r: -(_num(r.get("PAYMENTS_VOLUME_USD"))),
    )),
    ("Paying Total", dict(
        note="Ever had a subscription (SUBSCRIPTION_STATUS populated). "
             "BEST FIT: 2,540 rows vs 1,954 in the hand-built 'Paying Total_Since 2025'.",
        rule=lambda r: str(r.get("SUBSCRIPTION_STATUS") or "").strip() not in ("", "None"),
        sort=lambda r: -(_num(r.get("POS_GTV_12M_USD"))),
    )),
])


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ---------------------------------------------------------------------------
def style(cell, kind):
    if kind == "title":
        cell.font = TITLE_F
    elif kind == "head":
        cell.font = HDR_F
        cell.fill = PatternFill("solid", fgColor=NAVY)
    elif kind == "colhead":
        cell.font = HDR_F
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")
    elif kind == "note":
        cell.font = NOTE_F
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    elif kind.startswith("input"):
        cell.font = INPUT_F
        cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
        cell.number_format = {"input_pct": PCT4, "input_money": MONEY,
                              "input_int": INT}.get(kind, "General")
    elif kind == "money":
        cell.font = BODY_F
        cell.number_format = MONEY
    elif kind == "int":
        cell.font = BODY_F
        cell.number_format = INT
    elif kind in ("pct", "pctlabel"):
        cell.font = BODY_F
        cell.number_format = PCT4
    else:
        cell.font = BODY_F


def build_assumptions(wb, carry, stats):
    if "Margin Assumptions" in wb.sheetnames:
        del wb["Margin Assumptions"]
    ws = wb.create_sheet("Margin Assumptions")
    ws.column_dimensions["A"].width = 58
    for c in "BCD":
        ws.column_dimensions[c].width = 18

    for r, c, v, kind in ASSUMPTIONS:
        ref = f"{gcl(c)}{r}"
        if kind.startswith("input") and ref in carry:
            v = carry[ref]
        cell = ws.cell(r, c, v)
        style(cell, kind)
        # Only the prose notes in column A get merged across. The tier band labels on
        # row 19 are also kind 'note' but sit in B/C/D; merging B19:F19 first would
        # make C19 a read-only MergedCell and the next write would fail.
        if kind == "note" and c == 1:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 26

    # Ranges are built from the ACTUAL column letters the model landed on, not from
    # the MODEL dict keys. The keys are the letters those columns happened to have in
    # the hand-built workbook; if the generated column count ever changes, the model
    # block shifts and using the keys here would point every total at the wrong data.
    def _r(letter):
        return f"'{DATA_SHEET}'!${letter}${DATA_START}:${letter}${MAXROW}"

    rngs = {f"{k}_R": _r(letter) for k, letter in stats["layout"].items()
            if MODEL[k][0]}
    rngs["GTV_R"] = _r(stats["gtv_col"])
    for r, c, v, kind in OUTPUTS:
        val = v.format(**rngs) if isinstance(v, str) and "{" in v else v
        style(ws.cell(r, c, val), kind)
    return ws


class PreviousWorkbookUnreadable(RuntimeError):
    """The previous workbook exists but could not be opened."""


def _open_previous(prev_path, **kw):
    """Open the previous published workbook.

    A MISSING file is legitimate - that is the first run. An UNREADABLE file is not,
    and must never be treated as 'no previous values'. Silently falling back to the
    ASSUMPTIONS defaults would republish the 21 Aug prices, which look entirely
    plausible, and quietly discard whatever Felipe had edited since. Verified: a
    torn or zero-byte workbook returned 0 carried cells with only a note in the log.

    A file is unreadable mainly when something is mid-write: a half-finished copy, or
    Excel partway through saving over the local mirror.
    """
    if not prev_path or not pathlib.Path(prev_path).exists():
        return None
    try:
        return openpyxl.load_workbook(prev_path, **kw)
    except Exception as e:
        raise PreviousWorkbookUnreadable(
            f"{prev_path} exists but could not be opened ({e}). Refusing to continue: "
            "carrying on would republish default assumptions and drop hand-typed "
            "ACTION text. If the file is open in Excel or being synced, wait for the "
            "write to finish and re-run."
        ) from e


def carry_inputs(prev_path):
    """Read the previous workbook's Margin Assumptions inputs so a price Felipe
    edited in Excel survives the next rebuild."""
    wb = _open_previous(prev_path, read_only=True, data_only=False)
    if wb is None:
        return {}
    if "Margin Assumptions" not in wb.sheetnames:
        return {}
    ws = wb["Margin Assumptions"]
    out = {}
    for ref in INPUT_CELLS:
        v = ws[ref].value
        if v is not None and not (isinstance(v, str) and v.startswith("=")):
            out[ref] = v
    wb.close()
    return out


def seed_action(path):
    """One-time seed of the hand-typed ACTION column.

    ACTION exists only because Felipe typed in it, and the FIRST automated run has no
    previous published workbook to merge from - so without this the eight entries in
    the 21 Aug workbook would be lost on run one. Recovered by streaming the sheet XML
    (openpyxl runs out of memory on that 496 MB sheet) and committed as
    data/action_seed.csv. After the first successful run the previous workbook takes
    over and this file is only a backstop.
    """
    import csv as _csv
    if not path or not pathlib.Path(path).exists():
        return {}
    out = {}
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in _csv.DictReader(fh):
            mid, ac = (r.get("MERCHANT_ID") or "").strip(), (r.get("ACTION") or "").strip()
            if mid and ac:
                out[mid] = ac
    return out


def carry_action(prev_path, id_header="MERCHANT_ID", action_header="ACTION"):
    """ACTION is hand-typed and is NOT in the SQL - Q1's SELECT ends at PULLED_AT.
    A full rebuild would erase it, so it is merged forward on MERCHANT_ID."""
    wb = _open_previous(prev_path, read_only=True, data_only=True)
    if wb is None:
        return {}
    if DATA_SHEET not in wb.sheetnames:
        return {}
    ws = wb[DATA_SHEET]
    it = ws.iter_rows(min_row=1, max_row=4)
    hdr = None
    for row in it:
        vals = [c.value for c in row]
        if id_header in vals:
            hdr = vals
            break
    if not hdr or action_header not in hdr:
        wb.close()
        return {}
    i_id, i_ac = hdr.index(id_header), hdr.index(action_header)
    out = {}
    for row in ws.iter_rows(min_row=2):
        if i_id >= len(row) or i_ac >= len(row):
            continue
        mid, ac = row[i_id].value, row[i_ac].value
        if mid is not None and ac not in (None, ""):
            out[str(mid)] = ac
    wb.close()
    return out


def read_headers(ws):
    return [c.value for c in next(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW))]


def add_model_columns(ws, headers, prev_actions):
    need = ("POS_GTV_12M_USD", "POS_RECEIPTS_12M", "MERCHANT_ID")
    for n in need:
        if n not in headers:
            sys.exit(f"Full Base is missing required column {n!r}; cannot build the model")
    gtv = gcl(headers.index("POS_GTV_12M_USD") + 1)
    rcp = gcl(headers.index("POS_RECEIPTS_12M") + 1)
    i_id = headers.index("MERCHANT_ID")

    base = len(headers)
    if "ACTION" not in headers:
        base += 1
        a_col = base
        c = ws.cell(HDR_ROW, a_col, "ACTION")
        c.font, c.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
        ws.column_dimensions[gcl(a_col)].width = 34
    else:
        a_col = headers.index("ACTION") + 1
    base += 1                       # one spacer column before the model block

    start = base + 1
    layout = {}
    for i, key in enumerate(MODEL):
        layout[key] = start + i

    for key, col in layout.items():
        header, fmt, _f = MODEL[key]
        if header is None:
            continue
        c = ws.cell(HDR_ROW, col, header)
        c.font, c.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[gcl(col)].width = 17
    ws.row_dimensions[HDR_ROW].height = 30

    # Formula columns reference each other by the MODEL keys (CI, CL, ...), which are
    # the letters those columns had in the hand-built workbook. Remap to wherever they
    # actually land now.
    remap = {k: gcl(v) for k, v in layout.items()}

    n_rows = n_formula = n_action = 0
    last = ws.max_row
    for r in range(DATA_START, last + 1):
        mid = ws.cell(r, i_id + 1).value
        if mid is None:
            continue
        n_rows += 1

        if prev_actions:
            a = prev_actions.get(str(mid))
            if a:
                ws.cell(r, a_col).value = a
                n_action += 1

        # ONLY where there is GTV. 88.8% of rows have none, and for those every model
        # formula returns "" - see design note 1.
        if _num(ws.cell(r, headers.index("POS_GTV_12M_USD") + 1).value) <= 0:
            continue

        for key, col in layout.items():
            header, fmt, f = MODEL[key]
            if f is None:
                continue
            expr = f.format(G=gtv, R=rcp, row=r)
            for k, letter in remap.items():
                if k != letter:
                    expr = expr.replace(f"${k}{r}", f"${letter}{r}")
            cell = ws.cell(r, col, expr)
            cell.font = BODY_F
            if fmt:
                cell.number_format = fmt
            n_formula += 1

    return dict(rows=n_rows, formulas=n_formula, actions=n_action,
                gtv_col=gtv, layout=remap)


def build_snapshots(wb, headers):
    src = wb[DATA_SHEET]
    i_id = headers.index("MERCHANT_ID")
    keep = [h for h in headers if h]
    idx = {h: headers.index(h) for h in keep}

    rows = []
    for row in src.iter_rows(min_row=DATA_START):
        if i_id >= len(row) or row[i_id].value is None:
            continue
        rows.append({h: row[idx[h]].value if idx[h] < len(row) else None for h in keep})

    counts = {}
    for name, cfg in SNAPSHOTS.items():
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws["A1"] = f"{name} — {cfg['note']}"
        ws["A1"].font = NOTE_F
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        ws.row_dimensions[1].height = 28

        for i, h in enumerate(keep, start=1):
            c = ws.cell(2, i, h)
            c.font, c.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
        sel = [r for r in rows if cfg["rule"](r)]
        try:
            sel.sort(key=cfg["sort"])
        except TypeError:
            pass
        for j, r in enumerate(sel):
            for i, h in enumerate(keep, start=1):
                ws.cell(3 + j, i, r[h])
        ws.freeze_panes = "C3"
        counts[name] = len(sel)

    # EMAIL BASE: one row per contactable email, tagged with the highest-priority
    # snapshot it belongs to.
    if "EMAIL BASE" in wb.sheetnames:
        del wb["EMAIL BASE"]
    ws = wb.create_sheet("EMAIL BASE")
    for i, h in enumerate(["EMAIL", "BASE", "GTV L12M"], start=1):
        c = ws.cell(1, i, h)
        c.font, c.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
    seen = {}
    for name, cfg in SNAPSHOTS.items():
        for r in rows:
            e = r.get("EMAIL")
            if e and e not in seen and cfg["rule"](r):
                seen[e] = (name, _num(r.get("POS_GTV_12M_USD")))
    for j, (e, (nm, g)) in enumerate(sorted(seen.items(), key=lambda kv: -kv[1][1])):
        ws.cell(2 + j, 1, e)
        ws.cell(2 + j, 2, nm)
        ws.cell(2 + j, 3, g).number_format = MONEY
    counts["EMAIL BASE"] = len(seen)
    return counts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--previous", default=None)
    ap.add_argument("--action-seed", default=None,
                    help="CSV backstop for the hand-typed ACTION column; "
                         "used only for entries the previous workbook lacks")
    a = ap.parse_args()

    print(f"reading {a.workbook}")
    wb = openpyxl.load_workbook(a.workbook, data_only=False)
    if DATA_SHEET not in wb.sheetnames:
        sys.exit(f"{a.workbook} has no {DATA_SHEET!r} sheet — run build_full_base.py first")

    carry = carry_inputs(a.previous)
    if carry:
        print(f"  carried {len(carry)} Margin Assumptions input(s) from the previous run")
    actions = seed_action(a.action_seed)
    if actions:
        print(f"  seeded {len(actions)} ACTION entr(ies) from {a.action_seed}")
    # The previous published workbook wins: it reflects anything typed since the seed.
    carried = carry_action(a.previous)
    if carried:
        print(f"  carried {len(carried)} hand-typed ACTION entr(ies) from the previous run")
    actions.update(carried)

    ws = wb[DATA_SHEET]
    headers = read_headers(ws)
    print(f"  {DATA_SHEET}: {ws.max_row - 1:,} data rows, {len(headers)} generated columns")

    stats = add_model_columns(ws, headers, actions)
    print(f"  model columns appended: {stats['formulas']:,} formula cells over "
          f"{stats['rows']:,} rows (GTV column {stats['gtv_col']})")
    print(f"  ACTION merged onto {stats['actions']} row(s)")

    build_assumptions(wb, carry, stats)
    print("  Margin Assumptions rebuilt (inputs blue, ranges bounded at "
          f"{MAXROW:,})")

    counts = build_snapshots(wb, headers)
    for k, v in counts.items():
        print(f"  snapshot {k:32} {v:>7,} rows")

    wb.calculation.fullCalcOnLoad = True
    wb.save(a.workbook)
    print(f"\nwrote {a.workbook}")


if __name__ == "__main__":
    main()
