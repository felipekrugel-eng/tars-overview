#!/usr/bin/env python3
"""
run.py — merchant-base-automation pipeline: pull -> build -> model -> recalc ->
verify -> publish.

    PULL      q1_base.sql against Snowflake  -> data/q1_export.csv
    BUILD     build_full_base.py             -> the 7 generated tabs
    MODEL     merchant_base_model.py         -> Felipe's margin model, Margin
                                                Assumptions, snapshot tabs, and the
                                                ACTION merge
    RECALC    LibreOffice headless
    VERIFY    the gates below
    PUBLISH   output/US_Merchant_Base_FULL.xlsx

Daily, not hourly. q2 hits 8.76bn receipt rows and takes 20-30 minutes; q1 alone is
minutes, but the merchant base does not move fast enough to justify more.

FALLBACK: if the pull fails its sanity gate (fewer than MIN_ROWS rows), the last
good committed CSV is reused so the workbook still refreshes on schedule - it just
will not carry same-day data. Same behaviour as payments-automation.

Env:
    SNOWFLAKE_*     connection (private key auth, same secret as the CASE/kpi pulls)
    SQL_DIR, DATA_DIR, OUT_DIR, TEMPLATE_PREV
    SKIP_PULL=1     build from the committed CSV without touching Snowflake
"""

import os
import pathlib
import shutil
import subprocess
import sys

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent
NAME = "US_Merchant_Base_FULL.xlsx"

SQL_DIR = pathlib.Path(os.environ.get("SQL_DIR", HERE / "queries"))
DATA_DIR = pathlib.Path(os.environ.get("DATA_DIR", HERE / "data"))
OUT_DIR = pathlib.Path(os.environ.get("OUT_DIR", HERE / "output"))
# The export is committed GZIPPED. Plain it is ~53 MB and this job commits it every
# day; git history is forever, so that alone was ~19 GB/year. gzip measures at 9.2% on
# real data (53.2 MB -> ~4.9 MB). build_full_base.py reads plain CSV and is left
# untouched, so run.py decompresses to a working copy that is gitignored.
CSV_GZ = DATA_DIR / "q1_export.csv.gz"      # committed
CSV = DATA_DIR / "q1_export.csv"            # working copy, NOT committed
OUT_XLSX = OUT_DIR / NAME
BUILD = HERE / "_build.xlsx"
RECALC_DIR = HERE / "_recalc"

# The previous published workbook. Read-only, and only for the two things that
# cannot be regenerated: the Margin Assumptions inputs and hand-typed ACTION text.
PREV = pathlib.Path(os.environ.get("TEMPLATE_PREV", OUT_XLSX))

MIN_ROWS = int(os.environ.get("MIN_ROWS", 100_000))   # US base ~131k; far below = bad pull
DATA_SHEET = "Full Base"


def gzip_to(src, dst):
    """Compress src -> dst DETERMINISTICALLY.

    Two things make a .gz differ for identical input, and both would make git see a
    change on every single run - re-committing ~5 MB daily even when the export is
    byte-identical, which defeats the reason for compressing at all:
      * the mtime in the gzip header  -> mtime=0
      * the original FILENAME in the header. GzipFile(path) embeds it, and passing a
        fileobj is NOT enough - gzip falls back to fileobj.name, so the destination
        path leaks into the header anyway. filename="" is what actually suppresses it.
    Verified: same input now yields byte-identical output.
    """
    import gzip
    with open(src, "rb") as fi, open(dst, "wb") as raw, \
            gzip.GzipFile(filename="", mode="wb", fileobj=raw, mtime=0) as fo:
        shutil.copyfileobj(fi, fo, length=1 << 20)


def gunzip_to(src, dst):
    import gzip
    with gzip.open(src, "rb") as fi, open(dst, "wb") as fo:
        shutil.copyfileobj(fi, fo, length=1 << 20)


def sh(*cmd, **kw):
    print("+ " + " ".join(str(c) for c in cmd), flush=True)
    subprocess.run([str(c) for c in cmd], check=True, **kw)


def publish(src, dst):
    """Replace dst with src atomically.

    On the CI runner nothing has the file open and a plain copy would do. This matters
    for a LOCAL MIRROR: shutil.copyfile opens the destination for writing and TRUNCATES
    it, so a reader that had it open sees a half-written file for the duration. Writing
    a sibling temp file and then os.replace() makes the swap a single atomic rename -
    a reader either gets the whole old file or the whole new one, never a torn one.

    An Excel lock file (~$name.xlsx) means someone has the workbook open. Publishing
    anyway is safe for the file on disk, but Excel still holds the OLD content in
    memory and saving from there would overwrite the fresh build with stale data. So
    skip, and say so, rather than set up a silent lost update.
    """
    dst = pathlib.Path(dst)
    lock = dst.with_name("~$" + dst.name)
    if lock.exists():
        print(f"SKIPPING PUBLISH: {dst.name} is open in Excel ({lock.name} present).\n"
              f"  The build is good and is in {src}. Close the workbook and re-run.\n"
              "  Publishing now would leave Excel holding stale content that a save "
              "would write back over this build.", file=sys.stderr)
        return False
    tmp = dst.with_name(dst.name + ".tmp")
    shutil.copyfile(src, tmp)
    os.replace(tmp, dst)
    return True


def soffice():
    for n in (os.environ.get("SOFFICE"), "soffice", "libreoffice"):
        if n and shutil.which(n):
            return shutil.which(n)
    sys.exit("no LibreOffice on PATH")


def pull():
    """Run q1_base.sql and write data/q1_export.csv. Keeps the last good CSV on
    failure rather than publishing a truncated base."""
    if os.environ.get("SKIP_PULL") == "1":
        print("SKIP_PULL=1 — using the committed CSV")
        return False
    try:
        import snowflake.connector
    except ImportError:
        print("snowflake-connector-python not installed — using the committed CSV")
        return False

    key = os.environ.get("SNOWFLAKE_PRIVATE_KEY")
    if not key:
        print("SNOWFLAKE_PRIVATE_KEY unset — using the committed CSV")
        return False

    import csv as _csv
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives import serialization

    pkey = serialization.load_pem_private_key(
        key.encode(), password=None, backend=default_backend()
    ).private_bytes(encoding=serialization.Encoding.DER,
                    format=serialization.PrivateFormat.PKCS8,
                    encryption_algorithm=serialization.NoEncryption())

    sql = (SQL_DIR / "q1_base.sql").read_text()
    tmp = DATA_DIR / "_q1_export.partial.csv"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = snowflake.connector.connect(
        account=os.environ["SNOWFLAKE_ACCOUNT"], user=os.environ["SNOWFLAKE_USER"],
        private_key=pkey, role=os.environ.get("SNOWFLAKE_ROLE", "DATA_VIEWER"),
        warehouse=os.environ.get("SNOWFLAKE_WAREHOUSE", "COMPUTE_WH"),
        database=os.environ.get("SNOWFLAKE_DATABASE", "LOYVERSE_DATA_LAKE"),
        schema=os.environ.get("SNOWFLAKE_SCHEMA", "PUBLIC"),
    )
    try:
        cur = conn.cursor()
        cur.execute("ALTER SESSION SET STATEMENT_TIMEOUT_IN_SECONDS = 3600")
        cur.execute(sql)
        cols = [d[0] for d in cur.description]
        n = 0
        with open(tmp, "w", newline="", encoding="utf-8") as fh:
            w = _csv.writer(fh)
            w.writerow(cols)
            while True:
                batch = cur.fetchmany(10_000)
                if not batch:
                    break
                for row in batch:
                    w.writerow(["" if v is None else v for v in row])
                    n += 1
        print(f"  pulled {n:,} rows")
        if n < MIN_ROWS:
            print(f"  SANITY GATE: {n:,} < {MIN_ROWS:,} — keeping the last good CSV",
                  file=sys.stderr)
            tmp.unlink(missing_ok=True)
            return False
        shutil.move(tmp, CSV)
        # Compress for the commit. mtime=0 so an unchanged export produces a
        # byte-identical .gz and git sees no diff - otherwise the gzip header
        # timestamp would make every run look like a change and defeat the point.
        gzip_to(CSV, CSV_GZ)
        print(f"  compressed -> {CSV_GZ.name} "
              f"({CSV_GZ.stat().st_size / 1e6:.1f} MB, "
              f"{CSV_GZ.stat().st_size / CSV.stat().st_size:.1%} of plain)")
        return True
    finally:
        conn.close()


def recalc(src):
    # ignore_errors: see margin-automation/run.py - an unguarded rmtree kills an
    # otherwise good run on any filesystem that refuses unlink.
    shutil.rmtree(RECALC_DIR, ignore_errors=True)
    RECALC_DIR.mkdir(parents=True, exist_ok=True)
    sh(soffice(), "--headless", "--norestore", "--convert-to", "xlsx",
       "--outdir", RECALC_DIR, src)
    out = RECALC_DIR / src.name
    if not out.exists():
        sys.exit(f"recalc produced nothing at {out}")
    return out


def verify(path, prev_actions):
    problems = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    names = set(wb.sheetnames)
    required = {"Read Me", "Base Summary", "Payments Funnel", "GTV & Opportunity",
                "By Cohort Month", "Data Coverage", DATA_SHEET, "Margin Assumptions",
                "Prime Base", "Transacting Merchants", "EMAIL BASE"}
    missing = required - names
    if missing:
        problems.append(f"sheets missing: {sorted(missing)}")

    # The junk tab must never come back. In the hand-built file it carried 918,622
    # row elements of which 1,294 held data - 730 MB of empty rows, and the single
    # largest object in a 151 MB workbook.
    if "PAYING_ACTIVE N_ACTIVE" in names:
        problems.append("'PAYING_ACTIVE N_ACTIVE' is present — that tab was 917,328 "
                        "empty rows and must not be regenerated")

    if DATA_SHEET in names:
        ws = wb[DATA_SHEET]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for need in ("MERCHANT_ID", "BASE_GROUP", "PRIORITY_BAND", "IS_TARGETABLE",
                     "IS_CONTACTABLE", "POS_GTV_12M_USD", "ACTION"):
            if need not in hdr:
                problems.append(f"{DATA_SHEET} is missing column {need!r}")
        i_id = hdr.index("MERCHANT_ID") if "MERCHANT_ID" in hdr else 0
        i_ac = hdr.index("ACTION") if "ACTION" in hdr else None
        n = 0
        kept = 0
        for row in ws.iter_rows(min_row=2):
            if i_id < len(row) and row[i_id].value is not None:
                n += 1
                if i_ac is not None and i_ac < len(row) and row[i_ac].value:
                    kept += 1
        if n < MIN_ROWS:
            problems.append(f"{DATA_SHEET} has {n:,} rows, expected >= {MIN_ROWS:,}")
        # ACTION is hand-typed and not in the SQL. Losing it is silent data loss.
        if prev_actions and kept < prev_actions:
            problems.append(f"ACTION entries dropped: {kept} kept of {prev_actions} "
                            "carried forward")
        print(f"VERIFY: rows={n:,} action_entries={kept} sheets={len(names)}", flush=True)

    errs = {}
    for ws in wb.worksheets:
        if ws.title == DATA_SHEET:
            continue        # 350k formulas; the recalc log already surfaces failures
        c = sum(1 for row in ws.iter_rows() for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("#"))
        if c:
            errs[ws.title] = c
    if errs:
        problems.append(f"formula errors: {errs}")

    wb.close()
    return problems


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    print("=== PULL ===", flush=True)
    fresh = pull()
    if not CSV.exists() and CSV_GZ.exists():
        # Pull skipped or failed its sanity gate: fall back to the last good committed
        # export by decompressing it.
        print(f"  decompressing {CSV_GZ.name} (last good export)")
        gunzip_to(CSV_GZ, CSV)
    if not CSV.exists():
        sys.exit(f"no export at {CSV_GZ} or {CSV}, and the pull did not produce one")
    print(f"  using {CSV} ({CSV.stat().st_size / 1e6:.1f} MB, "
          f"{'fresh' if fresh else 'last good'})")

    # Establish the ACTION baseline BEFORE building, and make it non-vacuous. The
    # earlier version computed it with the same carry_action() that the model step
    # uses, so an unreadable previous workbook gave 0 here AND 0 there - and the
    # verify check "kept < prev_actions" passed trivially while the entries were lost.
    # seed_action is a plain CSV and cannot fail this way, so it floors the baseline.
    from merchant_base_model import (PreviousWorkbookUnreadable, carry_action,
                                     seed_action)
    seeded = seed_action(DATA_DIR / "action_seed.csv")
    try:
        carried = carry_action(PREV)
    except PreviousWorkbookUnreadable as e:
        sys.exit(f"ABORTING BEFORE BUILD: {e}")
    prev_actions = len({**seeded, **carried})
    print(f"  ACTION baseline: {len(seeded)} seeded + {len(carried)} carried "
          f"= {prev_actions} expected to survive")

    print("\n=== BUILD ===", flush=True)
    sh(sys.executable, HERE / "build_full_base.py", CSV, BUILD)

    print("\n=== MODEL ===", flush=True)
    cmd = [sys.executable, HERE / "merchant_base_model.py", "--workbook", BUILD]
    seed = DATA_DIR / "action_seed.csv"
    if seed.exists():
        cmd += ["--action-seed", seed]
    if PREV.exists():
        cmd += ["--previous", PREV]
    sh(*cmd)

    print("\n=== RECALC ===", flush=True)
    recalced = recalc(BUILD)

    print("\n=== VERIFY ===", flush=True)
    problems = verify(recalced, prev_actions)
    if problems:
        print("VERIFY FAILED:", file=sys.stderr)
        for p in problems:
            print("  - " + p, file=sys.stderr)
        sys.exit(1)
    print("VERIFY OK", flush=True)

    print("\n=== PUBLISH ===", flush=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if not publish(recalced, OUT_XLSX):
        sys.exit(2)
    print(f"published {OUT_XLSX} ({OUT_XLSX.stat().st_size / 1e6:.1f} MB)")
    try:
        BUILD.unlink(missing_ok=True)
    except OSError as e:
        print(f"note: could not remove {BUILD.name} ({e.strerror}) — harmless")
    shutil.rmtree(RECALC_DIR, ignore_errors=True)


if __name__ == "__main__":
    main()
