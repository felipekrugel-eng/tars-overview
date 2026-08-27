#!/usr/bin/env python3
"""
build_full_base.py — turn the Q1 Snowflake export into the full US merchant base workbook.

    python3 build_full_base.py q1_export.csv "US_Merchant_Base_Full_18Aug.xlsx"

Scope: ALL ~131,217 live US merchants. Differentiation happens in columns
(BASE_GROUP / PRIORITY_BAND / IS_TARGETABLE / EXCLUSION_REASON), never by
dropping rows — a merchant missing from the file raises questions, a merchant
tagged 'dormant' answers them.

Design notes for a 131k-row workbook:
  * Written in openpyxl WRITE-ONLY mode and streamed straight from the CSV. A
    normal Workbook holds every Cell object in memory; at 131k x 47 that is
    ~6.2M cells and the process is killed. Nothing is ever fully materialised.
  * The Full Base sheet carries STATIC values and no formulas. The old file had
    47,844 in-sheet formulas over 15,948 rows; at 131k that becomes ~400k
    formulas and Excel stops being usable. Everything is computed in SQL instead.
  * The summary tabs DO carry real formulas (COUNTIF/COUNTIFS/SUM over Full Base)
    so they recalculate if the data sheet is refreshed in place.
  * Summary ranges are bounded at MAXROW, deliberately far above the row count,
    so appending rows later does not silently fall outside them. The old file's
    ranges stopped at row 15949, which is why the 2,276 new merchants would have
    been excluded from every total even after being pasted in.
  * openpyxl writes formulas with no cached values, and LibreOffice stalls
    recalculating a large sheet carrying an autofilter. So the file is written
    with fullCalcOnLoad and NO autofilter, converted by LibreOffice, and the
    autofilter is streamed back into the sheet XML afterwards.
"""

import csv
import datetime as dt
import os
import re
import shutil
import subprocess
import sys
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

# ---------------------------------------------------------------- config ----
MAXROW = 300_000          # formula range ceiling; ~2.3x the current base
DATA_SHEET = "Full Base"

NAVY = "0B2B3C"
BLUE = "1D8FE1"
WARN = "FFF4D6"
BAD = "FDECEC"
GREY = "F2F4F6"
GOOD = "EAF7EE"   # pale green, used for high-confidence link methods

TITLE_F = Font(name="Arial", size=13, bold=True, color="FFFFFF")
HDR_F = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_F = Font(name="Arial", size=10)
BOLD_F = Font(name="Arial", size=10, bold=True)
NOTE_F = Font(name="Arial", size=9, italic=True, color="5A6B75")

THIN = Side(style="thin", color="D5DBDF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

INT_COLS = {
    "TENURE_DAYS", "RECEIPTS_LIFETIME", "ACTIVE_MONTHS", "PAYMENTS_TXNS",
    "IS_ENTERED", "IS_SIGNED_UP", "IS_ENABLED", "IS_TRANSACTING",
    "REJECTED_BUT_TRANSACTING", "POS_RECEIPTS_12M", "POS_RECEIPTS_L30D",
}
MONEY_COLS = {
    "SOFTWARE_MRR", "PAYMENTS_VOLUME_USD", "PAYMENTS_REFUNDED_USD",
    "PAYMENTS_NET_USD", "PAYMENTS_L30D_USD", "POS_GTV_12M_USD",
    "POS_GTV_12M_UNCAPPED_USD", "POS_GTV_L30D_USD", "POS_MEDIAN_TICKET_USD",
    "GTV_OFF_PAYMENTS_L30D_USD",
}
PCT_COLS = {"PAYMENTS_ATTACH_RATE_L30D"}
# Plain numbers: neither money nor counts nor percentages. GTV_UNIT_SCALE is the
# 1.0-or-0.01 factor the query measured and applied; it must land as a number so
# a reader can filter on it rather than squint at left-aligned text.
NUM_COLS = {"GTV_UNIT_SCALE"}
DATE_COLS = {
    "JOINED_LOYVERSE", "SUBSCRIPTION_STARTED", "SUBSCRIPTION_CANCELLED",
    "JOINED_PAYMENTS", "FIRST_CHARGE_AT", "LAST_PAYMENTS_CHARGE", "PULLED_AT",
    "POS_LAST_RECEIPT_DATE",
}
WIDE = {"BUSINESS_NAME": 34, "EMAIL": 30, "ADDRESS": 30, "EXCLUSION_REASON": 38,
        "PRIORITY_BAND": 28, "CONTACT_CHANNEL": 16, "STRIPE_ACCOUNT_ID": 22,
        "SUBSCRIPTION_STATUS": 18, "FUNNEL_STAGE": 18, "CONTACT_NAME": 20,
        "DISABLED_REASON": 22, "GTV_DATA_FLAG": 44, "GTV_BAND": 24,
        "PAYMENTS_ATTACH_RATE_L30D": 16, "GTV_OFF_PAYMENTS_L30D_USD": 18,
        "POS_GTV_12M_UNCAPPED_USD": 18, "GTV_CONFIDENCE": 42}

# Reconciliation targets from window.__FUNNEL_BASES / __FUNNEL_STAGES_BY at
# 2026-08-18. Update when the pull date moves.
EXPECT = {"total": 131_217, "signed_up": 682, "enabled": 202, "transacting": 52,
          "rejected_txn": 2}
PREV_SCOPE = 15_948


# ------------------------------------------------------------- utilities ----
def cell(ws, value, *, font=BODY_F, fill=None, fmt=None, align=None,
         wrap=False, border=True):
    c = WriteOnlyCell(ws, value=value)
    c.font = font
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if align or wrap:
        c.alignment = Alignment(horizontal=align or "general", vertical="center",
                                wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def banner(ws, text, span):
    """Title row. Write-only mode has no merged cells, so the bar is painted."""
    row = [cell(ws, text, font=TITLE_F, fill=NAVY, align="left", border=False)]
    row += [cell(ws, None, fill=NAVY, border=False) for _ in range(span - 1)]
    ws.append(row)


def head(ws, labels, fill=BLUE):
    ws.append([cell(ws, l, font=HDR_F, fill=fill, align="center", wrap=True)
               for l in labels])


def note(ws, text, span):
    row = [cell(ws, text, font=NOTE_F, align="left", wrap=True, border=False)]
    row += [cell(ws, None, border=False) for _ in range(span - 1)]
    ws.append(row)


def widths(ws, spec):
    for letter, w in spec.items():
        ws.column_dimensions[letter] = ColumnDimension(ws, index=letter, width=w)


def col_of(headers, name):
    return get_column_letter(headers.index(name) + 1)


def rng(headers, name):
    L = col_of(headers, name)
    return f"'{DATA_SHEET}'!${L}$2:${L}${MAXROW}"


# ------------------------------------------------------------ data sheet ----
def write_data_sheet(wb, csv_path):
    """Stream the CSV into the sheet. Returns (headers, nrows, months)."""
    ws = wb.create_sheet(DATA_SHEET)

    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        rdr = csv.reader(fh)
        headers = [h.strip().upper() for h in next(rdr)]

        required = ["MERCHANT_ID", "BUSINESS_NAME", "MONTH_JOINED_POS", "BASE_GROUP",
                    "PRIORITY_BAND", "IS_TARGETABLE", "IS_CONTACTABLE"]
        missing = [c for c in required if c not in headers]
        if missing:
            sys.exit(f"CSV is missing required columns: {missing}")

        # GTV_CONFIDENCE is derived here rather than in SQL because it needs the
        # Stripe cross-check, and re-running the query costs minutes. The rule is
        # documented in Read Me and reproduced in the master .sql for the next
        # pull, so the file and the query stay in step.
        gtv_conf = "POS_GTV_L30D_USD" in headers
        src = {}
        if gtv_conf:
            for c in ("POS_GTV_L30D_USD", "POS_MEDIAN_TICKET_USD", "GTV_DATA_FLAG",
                      "IS_TRANSACTING", "PAYMENTS_TXNS", "PAYMENTS_VOLUME_USD"):
                src[c] = headers.index(c) if c in headers else None
            headers = headers + ["GTV_CONFIDENCE"]

        spec = {}
        for i, h in enumerate(headers, start=1):
            L = get_column_letter(i)
            cd = ColumnDimension(ws, index=L, width=WIDE.get(h, max(11, min(20, len(h) + 3))))
            if h in MONEY_COLS:
                cd.number_format = '$#,##0.00;($#,##0.00);-'
            elif h in PCT_COLS:
                cd.number_format = '0.0%;-0.0%;-'
            elif h in INT_COLS:
                cd.number_format = '#,##0;(#,##0);-'
            elif h in NUM_COLS:
                cd.number_format = '0.00'
            elif h in DATE_COLS:
                cd.number_format = 'yyyy-mm-dd'
            ws.column_dimensions[L] = cd
            spec[i - 1] = h

        ws.freeze_panes = "C2"
        ws.append([cell(ws, h, font=HDR_F, fill=NAVY, align="center", wrap=True)
                   for h in headers])

        num_idx = {i for i, h in spec.items()
                   if h in INT_COLS or h in MONEY_COLS or h in PCT_COLS
                   or h in NUM_COLS}
        date_idx = {i for i, h in spec.items() if h in DATE_COLS}
        month_idx = headers.index("MONTH_JOINED_POS")

        months, n = set(), 0
        for rec in rdr:
            out = []
            for i, v in enumerate(rec):
                if v == "":
                    out.append(None)
                elif i in num_idx:
                    try:
                        f = float(v)
                        out.append(int(f) if f.is_integer() else f)
                    except ValueError:
                        out.append(v)
                elif i in date_idx:
                    out.append(_parse_dt(v))
                elif v[0] == "=":
                    # A merchant is literally named '=========='. openpyxl types
                    # any string starting with '=' as a formula, so it was written
                    # as one and came back #VALUE! after recalculation. Forcing
                    # the string type keeps the text verbatim and keeps the file
                    # free of formula errors.
                    c = WriteOnlyCell(ws, value=v)
                    c.data_type = "s"
                    c.font = BODY_F
                    out.append(c)
                else:
                    out.append(v)
            if gtv_conf:
                out.append(_confidence(rec, src))
            ws.append(out)
            if month_idx < len(rec) and rec[month_idx]:
                months.add(rec[month_idx])
            n += 1
            if n % 25_000 == 0:
                print(f"  ... {n:,} rows", flush=True)

    return headers, n, sorted(months)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# Volume above which a merchant's till total stops being credible on its own.
# $250k in 30 days is roughly $3m a year through a Loyverse till. Real merchants
# do reach it, which is why this DOWNGRADES rather than excludes, and why card
# corroboration overrides it entirely.
BIG_30D = 250_000


def _confidence(rec, src):
    """Tier a merchant's till volume by how much the number can be trusted.

    Measured on this export: across the 52 merchants who also process on Stripe
    -- the only place we hold volume in known-real dollars -- the POS median
    ticket tracks the real card ticket at 0.83x. So TOTAL_MONEY is in DOLLARS,
    settling the contradiction between us_paying_merchants_gtv.sql (dollars) and
    the dashboard's monthly.sql (cents). The scale is not the problem.

    What IS the problem is currency labelling: 555 merchants report USD while
    posting median tickets in the thousands, and they carry 89% of apparent GTV.
    Their names (Bodeguita, Ferreteria, ... SURL) point at local-currency tills
    tagged USD. Nobody is dropped -- per the standing rule the judgement goes in
    a column so it can be argued with, and the summary tabs total the tiers
    separately instead of publishing one contaminated headline.
    """
    g30 = _num(rec[src["POS_GTV_L30D_USD"]])
    if g30 <= 0:
        return "No recent till volume"

    med = _num(rec[src["POS_MEDIAN_TICKET_USD"]])
    txns = _num(rec[src["PAYMENTS_TXNS"]]) if src["PAYMENTS_TXNS"] is not None else 0
    vol = _num(rec[src["PAYMENTS_VOLUME_USD"]]) if src["PAYMENTS_VOLUME_USD"] is not None else 0

    # Card corroboration is the strongest evidence available and outranks both
    # heuristics: if a merchant's POS ticket matches the ticket Stripe actually
    # settled, the till is denominated in dollars whatever its size or average.
    # At least 5 charges, because a lone $1 test charge is not evidence.
    if txns >= 5 and vol > 0:
        stripe_avg = vol / txns
        if stripe_avg > 0 and med > 0 and 0.2 <= med / stripe_avg <= 5:
            return "1 - Confirmed by card data"

    flag = rec[src["GTV_DATA_FLAG"]] if src["GTV_DATA_FLAG"] is not None else ""
    if flag.startswith("Check"):
        return "3 - Check: ticket size suggests another currency"
    if g30 > BIG_30D:
        return "4 - Check: volume too large to take on trust"
    return "2 - Plausible"


def _parse_dt(v):
    v = v.strip().replace("T", " ")
    if v.endswith("Z"):
        v = v[:-1]
    v = re.sub(r"[+-]\d{2}:?\d{2}$", "", v).strip()
    for f in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = dt.datetime.strptime(v, f)
        except ValueError:
            continue
        # A date-only source string becomes a date, not a midnight datetime.
        # openpyxl formats date as 'yyyy-mm-dd' but datetime as
        # 'yyyy-mm-dd h:mm:ss', and that cell-level format beats the column
        # default -- which is why every date column was rendering a dead
        # ' 00:00:00'. Timestamps that really carry a time are left alone.
        return parsed.date() if f == "%Y-%m-%d" else parsed
    return v


# ---------------------------------------------------------- summary tabs ----
READ_ME = """
Scope: every live US Loyverse merchant. {nrows:,} rows, {ncols} columns.

WHAT CHANGED IN THIS VERSION

1. Scope is now the FULL US base, not a POS-active subset. The previous file held 15,948 rows,
   which was "Active on POS = Yes" as at a 12 July database cut. That silently excluded 2,276
   merchants who registered after the cut-off, and it implied a denominator of 15,948 when the
   true US base is 131,217. Every merchant now appears; targeting is expressed in columns.

2. Differentiation happens in BASE_GROUP, PRIORITY_BAND, IS_TARGETABLE and EXCLUSION_REASON.
   Filter, do not delete. A merchant absent from the file raises questions; a merchant tagged
   "dormant" with a stated reason answers them.

3. GTV_CONFIDENCE is new, and it is the most important thing in this file. Summing the till
   volume raw gives $619.8m over 30 days. That number is wrong, and the column shows why.
   555 merchants report CURRENCY = 'USD' while posting median tickets of $2,000-$34,000, and
   they carry 89% of that total on their own. Their names (Bodeguita, Ferreteria, ... SURL)
   and their addresses - 23 of the 555 have a US state on file - say these are local-currency
   tills labelled USD. The credible 30-day figure is $48.9m across 3,530 merchants, of which
   $48.7m is off our rails. That is the campaign, and it is the number to quote.

   Calibration, measured on this export rather than assumed: across the 52 merchants who also
   process on Stripe - the only place we hold volume in known-real dollars - the POS median
   ticket tracks the settled card ticket at 0.83x. So TOTAL_MONEY is in DOLLARS, which settles
   the standing disagreement between us_paying_merchants_gtv.sql (reads dollars) and the
   dashboard's monthly.sql (divides by minor units). The unit is not the problem. Currency
   labelling is.

   The tiers, in the order they are tested:
     1 - Confirmed by card data   POS ticket matches the ticket Stripe actually settled
                                  (>=5 charges, ratio within 0.2x-5x). Proven dollars, and
                                  this outranks both heuristics below - a real high-ticket
                                  US merchant is not demoted for having a large average.
     2 - Plausible                Ordinary ticket, ordinary volume, no reason to doubt it.
     3 - Check: ticket size       Median ticket over $200. Usually a non-USD till.
     4 - Check: volume            Over $250k in 30 days with no card data to corroborate it.
     No recent till volume        No clean USD receipts in the window. The dormant majority.

   Nobody is dropped, per the standing rule: the judgement sits in a column where it can be
   argued with, rather than being applied silently by leaving rows out. GTV & Opportunity
   totals the credible tiers in green and every row in amber, side by side, so the gap is
   visible instead of hidden. GTV_CONFIDENCE is derived in build_full_base.py, not in the
   query, because it needs the Stripe cross-check; the same rule is reproduced in the master
   .sql so the next pull computes it natively.

4. MONTH_JOINED_POS (the old column AA) is populated for 100% of rows, from
   LOYVERSE_MERCHANTS.CREATED_AT. It was entirely empty before, and only ~9% of it could be
   reconstructed from the funnel data feed.

5. EMAIL now comes from LOYVERSE_MERCHANTS rather than the Stripe connected account. Contact
   coverage was the real blocker on the previous file: email 4.0%, phone 1.0%, name 0.7%.
   A campaign list you cannot contact is not a campaign list.

6. No formulas on the Full Base sheet. Everything is computed in SQL and lands as a value.
   At this row count in-sheet formulas make the workbook unusable. Summary tabs still carry
   live formulas and recalculate.

COLUMN DEFINITIONS THAT MATTER

BASE_GROUP      Mutually exclusive, reproduces window.__FUNNEL_BASES exactly.
                new       - registered on or after the 2026-07-01 payments launch
                paying    - pre-launch, active Chargebee subscription
                nonpaying - pre-launch, no subscription, >=1 receipt in trailing 30d
                dormant   - pre-launch, no subscription, no receipt in trailing 30d

FUNNEL_STAGE    Mutually exclusive: one merchant, one stage. NOT the same shape as the
                dashboard, which is cumulative. Exclusive "Enabled" reads 150, not 202,
                because the 52 transacting merchants move up a stage. Use IS_SIGNED_UP /
                IS_ENABLED / IS_TRANSACTING to reconcile; use FUNNEL_STAGE to segment.

IS_TARGETABLE   "No" when the merchant is already processing payments, has no contact
                details, or is dormant. EXCLUSION_REASON states which.

KNOWN CAVEATS

* CREATED_AT is TIMESTAMP_LTZ but holds local wall-clock, not a true UTC instant. Safe for a
  YYYY-MM bucket; wrong for anything hourly.
* SOFTWARE_MRR grain is unverified - confirm major vs minor units before quoting revenue.
* CITY, STATE, ADDRESS, BUSINESS_TYPE, PHONE and CONTACT_NAME come from the Stripe connected
  account, so they exist only for merchants who started payments signup (~680). This is a real
  gap, not a pull error, and it caps what state-level segmentation can claim.
* REJECTED_BUT_TRANSACTING flags merchants whose KYC was rejected while money is still moving.
  Two existed at 18 Aug. Treat as a compliance item, not a data issue.

Source query: US_Merchant_Base_Refresh_Snowflake.sql (Q1). Rebuild: build_full_base.py.
""".strip("\n")


def build_read_me(wb, nrows, ncols):
    ws = wb.create_sheet("Read Me")
    widths(ws, {"A": 118})
    banner(ws, "US MERCHANT BASE - FULL DEFINITION OF CAMPAIGNS", 1)
    for line in READ_ME.format(nrows=nrows, ncols=ncols).split("\n"):
        bold = line and line == line.upper() and not line.startswith(("*", " "))
        ws.append([cell(ws, line, font=BOLD_F if bold else BODY_F,
                        align="left", border=False)])


def build_base_summary(wb, headers, nrows):
    ws = wb.create_sheet("Base Summary")
    widths(ws, {"A": 32, "B": 14, "C": 12, "D": 14, "E": 14, "F": 50})
    grp, tgt, con = (rng(headers, c) for c in ("BASE_GROUP", "IS_TARGETABLE", "IS_CONTACTABLE"))
    pb, ex = rng(headers, "PRIORITY_BAND"), rng(headers, "EXCLUSION_REASON")

    banner(ws, "BASE SUMMARY - every US merchant, grouped", 6)
    ws.append([])
    head(ws, ["Base group", "Merchants", "% of base", "Contactable", "Targetable", "What it means"])

    r = 4                                   # first data row
    groups = [
        ("new", "Registered since the 2026-07-01 payments launch"),
        ("paying", "Pre-launch, active subscription - the revenue base"),
        ("nonpaying", "Pre-launch, transacting on POS, not subscribed"),
        ("dormant", "Pre-launch, no subscription, no POS signal in 30 days"),
    ]
    total_row = r + len(groups)
    for i, (g, desc) in enumerate(groups):
        rr = r + i
        ws.append([
            cell(ws, g, font=BOLD_F),
            cell(ws, f'=COUNTIF({grp},$A{rr})', fmt='#,##0'),
            cell(ws, f'=IFERROR($B{rr}/$B${total_row},0)', fmt='0.0%'),
            cell(ws, f'=COUNTIFS({grp},$A{rr},{con},"Yes")', fmt='#,##0'),
            cell(ws, f'=COUNTIFS({grp},$A{rr},{tgt},"Yes")', fmt='#,##0'),
            cell(ws, desc),
        ])
    ws.append([
        cell(ws, "TOTAL", font=BOLD_F, fill=GREY),
        cell(ws, f'=SUM(B{r}:B{total_row-1})', font=BOLD_F, fill=GREY, fmt='#,##0'),
        cell(ws, 1.0, font=BOLD_F, fill=GREY, fmt='0.0%'),
        cell(ws, f'=SUM(D{r}:D{total_row-1})', font=BOLD_F, fill=GREY, fmt='#,##0'),
        cell(ws, f'=SUM(E{r}:E{total_row-1})', font=BOLD_F, fill=GREY, fmt='#,##0'),
        cell(ws, f"Must equal {EXPECT['total']:,} - the true US denominator", font=BOLD_F, fill=GREY),
    ])

    ws.append([])
    banner(ws, "PRIORITY BANDS - work order across the full base", 6)
    head(ws, ["Priority band", "Merchants", "% of base", "Contactable", "Targetable", "Note"])
    r = total_row + 4
    bands = [
        ("1 - Signed up, not processing", "Warmest: already in the payments funnel, stalled"),
        ("2 - Paying subscriber", "Proven willingness to pay Loyverse"),
        ("3 - Active, no subscription", "Transacting on POS, monetise"),
        ("4 - New registration", "Registered since launch, no history yet"),
        ("5 - Dormant", "Present for completeness, not for calling"),
    ]
    for i, (b, desc) in enumerate(bands):
        rr = r + i
        ws.append([
            cell(ws, b, font=BOLD_F),
            cell(ws, f'=COUNTIF({pb},$A{rr})', fmt='#,##0'),
            cell(ws, f'=IFERROR($B{rr}/$B${total_row},0)', fmt='0.0%'),
            cell(ws, f'=COUNTIFS({pb},$A{rr},{con},"Yes")', fmt='#,##0'),
            cell(ws, f'=COUNTIFS({pb},$A{rr},{tgt},"Yes")', fmt='#,##0'),
            cell(ws, desc),
        ])
    br = r + len(bands)
    ws.append([
        cell(ws, "TOTAL", font=BOLD_F, fill=GREY),
        cell(ws, f'=SUM(B{r}:B{br-1})', font=BOLD_F, fill=GREY, fmt='#,##0'),
        cell(ws, f'=SUM(C{r}:C{br-1})', font=BOLD_F, fill=GREY, fmt='0.0%'),
        cell(ws, f'=SUM(D{r}:D{br-1})', font=BOLD_F, fill=GREY, fmt='#,##0'),
        cell(ws, f'=SUM(E{r}:E{br-1})', font=BOLD_F, fill=GREY, fmt='#,##0'),
        cell(ws, "Bands are exhaustive - must also equal the base total", font=BOLD_F, fill=GREY),
    ])

    ws.append([])
    banner(ws, "WHY MERCHANTS ARE NOT TARGETED - stated, not hidden", 6)
    head(ws, ["Exclusion reason", "Merchants", "% of base", "", "", "Note"])
    r = br + 4
    reasons = [
        ("Dormant - no POS signal in 30d, no subscription", "Largest block by far; kept for denominator math"),
        ("No contact details", "Fixable - email is present on the merchant record"),
        ("Already processing payments", "Success, not a miss"),
    ]
    for i, (reason, desc) in enumerate(reasons):
        rr = r + i
        ws.append([
            cell(ws, reason),
            cell(ws, f'=COUNTIF({ex},$A{rr})', fmt='#,##0'),
            cell(ws, f'=IFERROR($B{rr}/$B${total_row},0)', fmt='0.0%'),
            cell(ws, None), cell(ws, None),
            cell(ws, desc),
        ])
    ws.append([])
    note(ws, f"Ranges are bounded at row {MAXROW:,}, well above the current {nrows:,} rows, so "
             f"appending merchants later stays inside every total. The previous workbook bounded "
             f"its ranges at row 15949, which is why newly added rows would have been silently "
             f"excluded from every summary.", 6)


def build_funnel(wb, headers):
    ws = wb.create_sheet("Payments Funnel")
    widths(ws, {"A": 32, "B": 14, "C": 14, "D": 12, "E": 54})
    banner(ws, "PAYMENTS FUNNEL - cumulative, reconciles to the dashboard", 5)
    ws.append([])
    head(ws, ["Stage", "Merchants", "Expected", "Delta", "Definition"])
    r = 4
    stages = [
        ("Entered", "IS_ENTERED", EXPECT["total"], "Every live US merchant in the base"),
        ("Signed up", "IS_SIGNED_UP", EXPECT["signed_up"], "Stripe connected account created"),
        ("Enabled", "IS_ENABLED", EXPECT["enabled"], "KYC passed, charges_enabled = TRUE"),
        ("Transacting", "IS_TRANSACTING", EXPECT["transacting"], "At least one succeeded, captured charge"),
    ]
    for i, (label, colname, expect, desc) in enumerate(stages):
        rr = r + i
        ws.append([
            cell(ws, label, font=BOLD_F),
            cell(ws, f'=SUM({rng(headers, colname)})', fmt='#,##0'),
            cell(ws, expect, fmt='#,##0'),
            cell(ws, f'=$B{rr}-$C{rr}', fmt='#,##0;(#,##0);-'),
            cell(ws, desc),
        ])
    ws.append([])
    note(ws, "'Expected' is window.__FUNNEL_STAGES_BY.US at 2026-08-18. A non-zero delta means "
             "the bot filter or the launch date has drifted - resolve that before trusting "
             "anything downstream. Entered is the full base here rather than the dashboard's "
             "3,721, which counts only post-launch entrants to the payments funnel.", 5)
    ws.append([])
    banner(ws, "COMPLIANCE FLAGS", 5)
    head(ws, ["Flag", "Merchants", "Expected", "Delta", "Action"])
    rr = r + len(stages) + 6
    ws.append([
        cell(ws, "KYC rejected, still transacting", font=BOLD_F),
        cell(ws, f'=SUM({rng(headers, "REJECTED_BUT_TRANSACTING")})', fmt='#,##0', fill=BAD),
        cell(ws, EXPECT["rejected_txn"], fmt='#,##0'),
        cell(ws, f'=$B{rr}-$C{rr}', fmt='#,##0;(#,##0);-'),
        cell(ws, "Filter Full Base on REJECTED_BUT_TRANSACTING = 1 and refer to compliance"),
    ])

    # Link quality. The Stripe share carries no Loyverse id, so every payments
    # row is an inference. This block says how strong each one is instead of
    # letting a reconstructed join pass for a native one.
    if "PAYMENTS_LINK_METHOD" in headers:
        ws.append([])
        ws.append([])
        banner(ws, "PAYMENTS LINK QUALITY - how each merchant was tied to its Stripe account", 5)
        head(ws, ["Link method", "Merchants", "Share of signed up", "", "How much to trust it"])
        lm = rng(headers, "PAYMENTS_LINK_METHOD")
        su = rng(headers, "IS_SIGNED_UP")
        methods = [
            ("Email (exact)", GOOD,
             "One Loyverse merchant holds that address. Solid."),
            ("Support email (exact)", GOOD,
             "Matched on the account's support address instead of its login address."),
            ("Email (shared address)", WARN,
             "Several merchants share the address - normally one owner, several stores. Spot-check."),
            ("Support email (shared)", WARN,
             "As above, matched via the support address."),
            ("Business name (exact)", WARN,
             "No email match; name is unique on both sides. Reasonable, not certain."),
            ("Manual override", GOOD,
             "Confirmed by a human and pinned in the query. Permanent."),
        ]
        for label, fill, desc in methods:
            ws.append([
                cell(ws, label),
                cell(ws, f'=COUNTIF({lm},"{label}")', fmt='#,##0', fill=fill),
                cell(ws, f'=IFERROR(COUNTIF({lm},"{label}")/SUM({su}),0)', fmt='0.0%'),
                cell(ws, ""),
                cell(ws, desc),
            ])
        ws.append([])
        note(ws, "The Stripe data share contains no Loyverse merchant id, so these links are "
                 "reconstructed from email and business name rather than read from a key. "
                 "Anything below 'exact' is a judgement call and is labelled as one. "
                 "Accounts that could not be matched confidently are left out of the base "
                 "entirely rather than attached to a guessed merchant - run "
                 "REVIEW_unlinked_stripe_accounts.sql to see them. The permanent fix is to "
                 "stamp the Loyverse id into Stripe account metadata at signup.", 5)


def build_cohorts(wb, headers, months):
    ws = wb.create_sheet("By Cohort Month")
    widths(ws, {"A": 16, "B": 14, "C": 12, "D": 12, "E": 13, "F": 13, "G": 14})
    banner(ws, "REGISTRATIONS BY MONTH - the old column AA, finally populated", 7)
    note(ws, "Sourced from LOYVERSE_MERCHANTS.CREATED_AT after the bot screen, so the April-July "
             "2026 bot campaign is already removed and this is the genuine registration curve.", 7)
    ws.append([])
    head(ws, ["Month joined", "Merchants", "New", "Paying", "Nonpaying", "Dormant", "Contactable"])
    mrng, grp, con = (rng(headers, c) for c in ("MONTH_JOINED_POS", "BASE_GROUP", "IS_CONTACTABLE"))
    r = 5
    for i, m in enumerate(months):
        rr = r + i
        row = [cell(ws, m, font=BOLD_F, align="center"),
               cell(ws, f'=COUNTIF({mrng},$A{rr})', fmt='#,##0')]
        for g in ("new", "paying", "nonpaying", "dormant"):
            row.append(cell(ws, f'=COUNTIFS({mrng},$A{rr},{grp},"{g}")', fmt='#,##0'))
        row.append(cell(ws, f'=COUNTIFS({mrng},$A{rr},{con},"Yes")', fmt='#,##0'))
        ws.append(row)
    last = r + len(months) - 1
    ws.append([cell(ws, "TOTAL", font=BOLD_F, fill=GREY)] +
              [cell(ws, f'=SUM({get_column_letter(c)}{r}:{get_column_letter(c)}{last})',
                    font=BOLD_F, fill=GREY, fmt='#,##0') for c in range(2, 8)])
    ws.freeze_panes = f"A{r}"


def build_gtv(wb, headers):
    """Till volume, attach rate, and the size of the prize.

    Only rendered when the export carries the GTV columns, so an older CSV still
    builds. Everything here is a live formula over Full Base rather than a value
    computed in Python: if the data sheet is refreshed in place, this tab moves
    with it instead of quietly describing last month's file.
    """
    if "POS_GTV_L30D_USD" not in headers:
        return

    ws = wb.create_sheet("GTV & Opportunity")
    widths(ws, {"A": 30, "B": 16, "C": 18, "D": 18, "E": 14, "F": 54})
    banner(ws, "TILL VOLUME vs PAYMENTS VOLUME - where the money actually is", 6)
    note(ws, "Every earlier version of this file counted receipts, never dollars. A merchant "
             "running $40k a month through somebody else's card reader looked identical to one "
             "running $400. These columns are the difference between a contact list and a "
             "prioritised one.", 6)
    ws.append([])

    gtv30 = rng(headers, "POS_GTV_L30D_USD")
    pay30 = rng(headers, "PAYMENTS_L30D_USD")
    opp = rng(headers, "GTV_OFF_PAYMENTS_L30D_USD")
    band = rng(headers, "GTV_BAND")
    flag = rng(headers, "GTV_DATA_FLAG")
    targ = rng(headers, "IS_TARGETABLE")

    conf = rng(headers, "GTV_CONFIDENCE") if "GTV_CONFIDENCE" in headers else None
    # Credible = tiers 1 and 2. SUMIF with a wildcard keeps this a live formula,
    # so it still holds if rows are filtered or the data sheet is refreshed.
    c1, c2 = '"1 - *"', '"2 - *"'

    def cred(rangetext):
        return (f'=SUMIF({conf},{c1},{rangetext})+SUMIF({conf},{c2},{rangetext})'
                if conf else f'=SUM({rangetext})')

    head(ws, ["Last 30 days", "Credible", "All rows", "", "", "What it means"])
    rows = [
        ("Till volume (POS GTV)", cred(gtv30), f'=SUM({gtv30})',
         "USD receipts across the base, cleaned, last 30 days"),
        ("Through Loyverse Payments", cred(pay30), f'=SUM({pay30})',
         "Succeeded, captured charges on connected accounts, same 30 days"),
        ("Still off our rails", cred(opp), f'=SUM({opp})',
         "THE PRIZE. Till volume we do not process today"),
    ]
    for label, cf, allf, desc in rows:
        ws.append([
            cell(ws, label, font=BOLD_F),
            cell(ws, cf, fmt='$#,##0;($#,##0);-', fill=GOOD),
            cell(ws, allf, fmt='$#,##0;($#,##0);-', fill=WARN),
            cell(ws, ""), cell(ws, ""),
            cell(ws, desc),
        ])
    ws.append([
        cell(ws, "Blended attach rate", font=BOLD_F),
        # Both sides need their own brackets. cred() returns "SUMIF(..)+SUMIF(..)",
        # and without parentheses round the numerator Excel binds the division to
        # the second SUMIF only -- which silently produced a dollar figure in a
        # percent-formatted cell instead of an attach rate.
        cell(ws, f'=IFERROR(({cred(pay30)[1:]})/({cred(gtv30)[1:]}),0)'
                 if conf else f'=IFERROR(SUM({pay30})/SUM({gtv30}),0)',
             fmt='0.00%', fill=GOOD),
        cell(ws, f'=IFERROR(SUM({pay30})/SUM({gtv30}),0)', fmt='0.00%', fill=WARN),
        cell(ws, ""), cell(ws, ""),
        cell(ws, "Payments volume as a share of till volume"),
    ])
    ws.append([])
    note(ws, "USE THE GREEN COLUMN. The amber column totals every row including tills that are "
             "almost certainly not denominated in dollars, and it overstates the opportunity by "
             "more than 12x. The split is explained in DATA QUALITY below; the per-merchant "
             "verdict is in GTV_CONFIDENCE on the data sheet.", 6)
    ws.append([])

    banner(ws, "BY SIZE BAND - who is worth an hour of somebody's time", 6)
    note(ws, "PRIORITY_BAND says who is warm. This says who is big. A dormant merchant doing "
             "$30k a month elsewhere outranks a warm one doing $200, and until now the file "
             "could not tell you which was which.", 6)
    head(ws, ["Size band (30d till volume)", "Merchants", "Till volume", "Off our rails",
              "Targetable", "Note"])
    bands = [
        ("A - $50k+ / 30d", "Rare, and each one is worth a personal call"),
        ("B - $20-50k / 30d", "Strong candidates, still small enough to move fast"),
        ("C - $5-20k / 30d", "The bulk of realistic near-term volume"),
        ("D - under $5k / 30d", "Long tail - campaign, do not call"),
        ("E - no recent till volume", "No POS signal in 30 days; GTV columns will read zero"),
    ]
    for label, desc in bands:
        q = f'"{label}"'
        ws.append([
            cell(ws, label, font=BOLD_F),
            cell(ws, f'=COUNTIF({band},{q})', fmt='#,##0'),
            cell(ws, f'=SUMIF({band},{q},{gtv30})', fmt='$#,##0;($#,##0);-'),
            cell(ws, f'=SUMIF({band},{q},{opp})', fmt='$#,##0;($#,##0);-'),
            cell(ws, f'=COUNTIFS({band},{q},{targ},"Yes")', fmt='#,##0'),
            cell(ws, desc),
        ])
    ws.append([])

    if conf:
        banner(ws, "HOW MUCH OF THIS CAN WE BELIEVE - confidence tiers", 6)
        note(ws, "Nobody is dropped. Each merchant is tiered by how much its till figure can be "
                 "trusted, so the judgement is visible in a column and can be argued with, "
                 "rather than being applied silently by leaving rows out.", 6)
        head(ws, ["Confidence", "Merchants", "Till volume", "Off our rails", "Targetable",
                  "Basis"])
        tiers = [
            ("1 - Confirmed by card data", GOOD,
             "Also processes on Stripe, and the POS ticket matches the ticket Stripe settled. "
             "Denominated in dollars, proven"),
            ("2 - Plausible", GOOD,
             "Ordinary ticket size and ordinary volume. No reason to doubt it"),
            ("3 - Check: ticket size suggests another currency", WARN,
             "Median ticket over $200. Usually a local-currency till labelled USD - the names "
             "skew heavily to non-US trading"),
            ("4 - Check: volume too large to take on trust", WARN,
             f"Over ${BIG_30D:,} in 30 days with no card data to corroborate it. Might be real, "
             "but confirm before it reaches a forecast"),
            ("No recent till volume", GREY,
             "No clean USD receipts in the last 30 days. The dormant majority"),
        ]
        for label, fill, desc in tiers:
            q = f'"{label}"'
            ws.append([
                cell(ws, label, font=BOLD_F),
                cell(ws, f'=COUNTIF({conf},{q})', fmt='#,##0', fill=fill),
                cell(ws, f'=SUMIF({conf},{q},{gtv30})', fmt='$#,##0;($#,##0);-'),
                cell(ws, f'=SUMIF({conf},{q},{opp})', fmt='$#,##0;($#,##0);-'),
                cell(ws, f'=COUNTIFS({conf},{q},{targ},"Yes")', fmt='#,##0'),
                cell(ws, desc),
            ])
        ws.append([])
        note(ws, "Calibration, measured on this export: across the 52 merchants who also process "
                 "on Stripe - the only place we hold volume in known-real dollars - the POS "
                 "median ticket tracks the settled card ticket at 0.83x. TOTAL_MONEY is "
                 "therefore in DOLLARS, which settles the long-running disagreement between "
                 "us_paying_merchants_gtv.sql (dollars) and the dashboard's monthly.sql (cents). "
                 "The unit is not the problem; currency labelling is.", 6)
        ws.append([])

    banner(ws, "DATA QUALITY - read this before quoting any number above", 6)
    head(ws, ["GTV data flag", "Merchants", "Till volume", "", "", "How to read it"])
    flags = [
        ("Clean", GOOD, "Cleaned USD receipts, nothing unusual"),
        ("Capped - one or more receipts over $10k", WARN,
         "Canonical $10k per-receipt cap applied. Compare against POS_GTV_12M_UNCAPPED_USD"),
        ("Check - median ticket over $200, likely non-USD data", WARN,
         "Canonical contamination signal: often a foreign-currency till labelled USD. "
         "Flagged, not dropped - decide per merchant"),
        ("No USD receipts in window", GREY,
         "No clean USD receipts in 12 months. Usually dormant, sometimes a non-USD till"),
    ]
    for label, fill, desc in flags:
        q = f'"{label}"'
        ws.append([
            cell(ws, label, font=BOLD_F),
            cell(ws, f'=COUNTIF({flag},{q})', fmt='#,##0', fill=fill),
            cell(ws, f'=SUMIF({flag},{q},{gtv30})', fmt='$#,##0;($#,##0);-'),
            cell(ws, ""), cell(ws, ""),
            cell(ws, desc),
        ])
    ws.append([])
    note(ws, "Unit safeguard: the query measures the median ticket rather than assuming a unit, "
             "and carries the factor it chose on every row in GTV_UNIT_SCALE - 1 means dollars, "
             "0.01 means cents. It read 1 on this pull, and the card-data calibration above "
             "independently confirms that. Check it again on each refresh.", 6)
    note(ws, "Windows: till volume is the last 30 calendar days; POS_GTV_12M_USD is the "
             "trailing 12 FULL months and deliberately excludes the current partial month. "
             "Attach rate divides 30 days by 30 days - never divide payments volume since "
             "the July launch by 12 months of GTV.", 6)


def build_coverage(wb, headers, nrows):
    ws = wb.create_sheet("Data Coverage")
    widths(ws, {"A": 30, "B": 14, "C": 12, "D": 66})
    banner(ws, "DATA COVERAGE - how full is each column, and why", 4)
    note(ws, "This tab exists because the previous file's gaps were invisible until audited: "
             "email 4.0%, phone 1.0%, contact name 0.7%, state 13.9%. Coverage is stated up "
             "front so nobody builds a campaign on a column that is mostly blank.", 4)
    ws.append([])
    head(ws, ["Column", "Populated", "% of rows", "Source and caveat"])
    src = [
        ("EMAIL", "LOYVERSE_MERCHANTS.EMAIL - present on essentially every registration"),
        ("PHONE", "Stripe connected account only - exists for payments signups (~680)"),
        ("CONTACT_NAME", "Stripe legal entity only - same limit as PHONE"),
        ("CITY", "Stripe legal entity address only"),
        ("STATE", "Stripe legal entity address only - caps state-level segmentation"),
        ("BUSINESS_TYPE", "Stripe connected account only"),
        ("MONTH_JOINED_POS", "LOYVERSE_MERCHANTS.CREATED_AT - must be 100%"),
        ("SOFTWARE_MRR", "CHARGEBEE_SUBSCRIPTIONS_V - subscribers only; grain unverified"),
        ("SUBSCRIPTION_STATUS", "CHARGEBEE_SUBSCRIPTIONS_V - subscribers only"),
        ("RECEIPTS_LIFETIME", "SALES_PER_ACCOUNT_MONTHLY - zero for merchants who never sold"),
        ("STRIPE_ACCOUNT_ID", "Stripe - payments signups only, by definition"),
        ("PAYMENTS_VOLUME_USD", "Stripe charges - transacting merchants only, by definition"),
        ("POS_GTV_12M_USD", "LOYVERSE_RECEIPTS - USD receipts, trailing 12 full months, cleaned"),
        ("POS_GTV_L30D_USD", "LOYVERSE_RECEIPTS - last 30 days; the number the attach rate uses"),
        ("PAYMENTS_ATTACH_RATE_L30D", "Blank where there was no till volume in 30d to divide by"),
        ("GTV_CONFIDENCE", "Derived in the build, not the query - see Read Me point 3"),
    ]
    r = 5
    n = 0
    for col, desc in src:
        if col not in headers:
            continue
        rr = r + n
        ws.append([
            cell(ws, col, font=BOLD_F),
            cell(ws, f'=COUNTA({rng(headers, col)})', fmt='#,##0'),
            cell(ws, f'=IFERROR($B{rr}/{nrows},0)', fmt='0.0%'),
            cell(ws, desc),
        ])
        n += 1
    ws.append([])
    banner(ws, "DENOMINATORS - quote these, never the row count of a filter", 4)
    head(ws, ["Denominator", "Merchants", "% of US base", "Note"])
    ws.append([
        cell(ws, "US base (this file)", font=BOLD_F),
        cell(ws, nrows, fmt='#,##0'),
        cell(ws, 1.0, fmt='0.0%'),
        cell(ws, "window.__FUNNEL_BASES.total_us at 2026-08-18"),
    ])
    ws.append([
        cell(ws, "Previous file's scope", font=BOLD_F, fill=WARN),
        cell(ws, PREV_SCOPE, fmt='#,##0', fill=WARN),
        cell(ws, PREV_SCOPE / nrows if nrows else 0, fmt='0.0%', fill=WARN),
        cell(ws, "POS-active as at a 12 July cut - a fraction of the base, presented as the whole",
             fill=WARN),
    ])


# ------------------------------------------------------- recalc pipeline ----
def recalc(path, filter_ref):
    """LibreOffice recalc, then stream the autofilter back into the sheet XML.

    The file is written with fullCalcOnLoad and no autofilter: openpyxl emits
    formulas with no cached values, and LibreOffice stalls recalculating a large
    sheet that carries a filter. Nothing is reloaded with openpyxl here - at this
    size that alone would exhaust memory.
    """
    outdir = os.path.join(os.path.dirname(os.path.abspath(path)), "_recalc")
    shutil.rmtree(outdir, ignore_errors=True)
    os.makedirs(outdir, exist_ok=True)
    subprocess.run(
        ["soffice", "--headless", "--norestore",
         "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", outdir, path],
        check=True, capture_output=True, timeout=2400,
    )
    produced = os.path.join(outdir, os.path.basename(path))
    if not os.path.exists(produced):
        raise RuntimeError("LibreOffice produced no output")
    _inject_filter(produced, DATA_SHEET, filter_ref)
    shutil.move(produced, path)
    shutil.rmtree(outdir, ignore_errors=True)


def _inject_filter(path, sheet_name, ref):
    """Insert <autoFilter/> after </sheetData>, streaming so the big sheet XML
    is never held in memory. CT_Worksheet requires that position."""
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    rid_target = dict(re.findall(r'Id="([^"]+)"[^>]*?Target="([^"]+)"', rels))
    target = None
    for m in re.finditer(r'<sheet[^>]*?name="([^"]+)"[^>]*?r:id="([^"]+)"', wbxml):
        if m.group(1) == sheet_name:
            target = "xl/" + rid_target.get(m.group(2), "").lstrip("/")
    if not target or target not in names:
        print(f"  ! could not locate {sheet_name} XML; autofilter not restored")
        return

    tag = f'<autoFilter ref="{ref}"/>'.encode()
    marker = b"</sheetData>"
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename != target:
                zout.writestr(item, zin.read(item.filename))
                continue
            with zin.open(item) as src, zout.open(item.filename, "w") as dst:
                done, buf = False, b""
                while True:
                    chunk = src.read(1 << 20)
                    if not chunk:
                        break
                    buf += chunk
                    if not done and marker in buf:
                        i = buf.index(marker) + len(marker)
                        dst.write(buf[:i] + tag)
                        buf = buf[i:]
                        done = True
                    if len(buf) > (1 << 21):
                        keep = len(marker)
                        dst.write(buf[:-keep])
                        buf = buf[-keep:]
                dst.write(buf)
                if not done:
                    print(f"  ! </sheetData> not found in {sheet_name}")
    shutil.move(tmp, path)


# ------------------------------------------------------------------ main ----
def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    csv_path = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "US_Merchant_Base_Full.xlsx"

    wb = Workbook(write_only=True)
    # Make Arial the workbook default so empty/unstyled cells don't fall back
    # to Calibri. Cheaper than styling 6M cells individually.
    try:
        normal = wb._named_styles["Normal"]
        normal.font.name = "Arial"
        normal.font.sz = 10
    except Exception:
        pass
    print("streaming data sheet...", flush=True)
    headers, nrows, months = write_data_sheet(wb, csv_path)
    print(f"  {nrows:,} rows x {len(headers)} columns", flush=True)

    build_read_me(wb, nrows, len(headers))
    build_base_summary(wb, headers, nrows)
    build_funnel(wb, headers)
    build_cohorts(wb, headers, months)
    build_gtv(wb, headers)
    build_coverage(wb, headers, nrows)

    order = ["Read Me", "Base Summary", "Payments Funnel", "GTV & Opportunity",
             "By Cohort Month", "Data Coverage", DATA_SHEET]
    wb._sheets = [wb[s] for s in order if s in wb.sheetnames]
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)
    print(f"wrote {out} ({os.path.getsize(out)/1e6:.1f} MB)", flush=True)

    filter_ref = f"A1:{get_column_letter(len(headers))}{nrows + 1}"
    print("recalculating (LibreOffice)...", flush=True)
    recalc(out, filter_ref)
    print(f"done: {out} ({os.path.getsize(out)/1e6:.1f} MB)")

    verify(out)


def verify(path):
    """Read back only the small sheets; never open Full Base."""
    wb = load_workbook(path, data_only=True, read_only=True)
    print("\nverification")
    for name in ("Base Summary", "Payments Funnel", "GTV & Opportunity", "Data Coverage"):
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=4, values_only=True):
            label, *vals = row
            if isinstance(label, str) and label.strip() and vals and vals[0] is not None:
                print(f"  {name:16s} | {str(label)[:44]:44s} | {vals[0]}")
    wb.close()


if __name__ == "__main__":
    main()
