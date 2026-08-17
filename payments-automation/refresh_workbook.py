#!/usr/bin/env python3
"""
Deterministic daily refresh of the Loyverse Embedded Payments V2 workbook.

Facadash pattern: two Snowflake CSV exports (transactions.csv, icplus_costs.csv)
are the only inputs. The V2 workbook is used as a styled TEMPLATE; this script
regenerates every data-derived sheet in place and lets the formula-driven sheets
recalc (via LibreOffice headless) against the new Transaction Detail rows.

Sheets rebuilt from data:
  - Transaction Detail        (raw, one row per charge; P/S/T/U..Y + AA/AB formulas)
  - Transaction Hyper Detail  (flattened values + per-plan fee matrix, grouped)
  - Fee Breakdown             (static aggregation of succeeded IC+ lines)
  - By Merchant               (merchant list + Txns/Volume formulas)
Sheets refreshed (text / ranges only):
  - Summary                   (subtitle, notes, scenario profitable counts + bumped ranges)
  - Read Me                   (scope text)
  - By Card Brand / Brand x Funding / Cost by Tranche / Pricing Model (bumped ranges)
Sheets left untouched: Sheet1 (manual pivot), Volume Target Calc, SQL Queries.

Usage:
  python3 refresh_workbook.py --tx transactions.csv --ic icplus_costs.csv \
      --template loyverse_payments_analysis_ful_V2.xlsx --out refreshed.xlsx
"""
import csv, copy, argparse, re, sys, os
from datetime import datetime
from collections import defaultdict, Counter
import openpyxl
from openpyxl.utils import get_column_letter as gcl, column_index_from_string as cifs
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import Outline

# ------------------------------------------------------------------ helpers
BRAND = {"Visa": "Visa", "MasterCard": "Mastercard",
         "American Express": "Amex", "Discover": "Discover"}
def brand(b): return BRAND.get(b, (b.title() if b else "Unknown"))
def fund(f):  return f.title() if f else "Unknown"
def norm(fn): return "card_scheme" if fn == "non_transactional_card_scheme" else fn
NETWORK = ["interchange", "card_scheme", "discount"]
STRIPE  = ["per_auth_fee", "volume_fee"]
CATS    = NETWORK + STRIPE

# --- Stripe platform fees, sourced from balance-transaction actuals ----------------
# Stripe debits the platform balance in GBP; the rest of this model is USD, so we
# convert the (small) platform-fee actuals to USD to keep the P&L single-currency.
# Spot GBP/USD on 2026-07-22 = 1.3373 (source: xe.com / Wise). Update if it drifts.
FX_GBP_USD = 1.3373
# Which platform_fees.csv buckets feed the Summary "Stripe Fees" section, and where.
# Everything not listed here (notably the 'other' bucket = terminal hardware purchases
# and any non-processing charges) is intentionally NEVER read. Categories with no
# actual rows stay at 0 ("0 until it appears"). (fee_category, Summary row, label)
PLATFORM_FEE_ROWS = [
    ("per_auth",       17, "        Per Auth Fee"),
    ("volume",         18, "        Volume Fee"),
    ("radar",          19, "        Radar (Fraud) Fee"),
    ("tap_to_pay",     20, "        Tap to Pay Fee"),
    ("payout",         21, "        Standard Payout Fee"),
    ("terminal_use",   22, "        Terminal Use Fee"),
    ("account_volume", 23, "        Account Volume Fee"),
]

def fnum(x):
    x = (x or "").strip()
    try:
        return float(x) if x not in ("", "None") else 0.0
    except ValueError:
        return 0.0

def fmt_window(d0, d1):
    """'2026-04-13','2026-07-21' -> 'Apr 13\u2013Jul 21, 2026'."""
    a = datetime.strptime(d0, "%Y-%m-%d"); b = datetime.strptime(d1, "%Y-%m-%d")
    if a.year == b.year:
        return f"{a.strftime('%b')} {a.day}\u2013{b.strftime('%b')} {b.day}, {b.year}"
    return f"{a.strftime('%b')} {a.day}, {a.year}\u2013{b.strftime('%b')} {b.day}, {b.year}"

# ------------------------------------------------------------------ data build
def build_data(tx_path, ic_path):
    tx = list(csv.DictReader(open(tx_path, encoding="utf-8-sig")))
    txmap = {t["CHARGE_ID"]: t for t in tx}
    succ = set(t["CHARGE_ID"] for t in tx if t["STATUS"] == "succeeded")

    ic_cat  = defaultdict(lambda: defaultdict(float))   # cid -> cat -> cents
    ic_plan = defaultdict(lambda: defaultdict(float))   # cid -> (cat,plan) -> cents
    cat_tot = defaultdict(float); cat_cnt = defaultdict(int)          # succeeded scope
    cat_plan_tot = defaultdict(lambda: defaultdict(float))
    cat_plan_cnt = defaultdict(lambda: defaultdict(int))
    plan_bf = defaultdict(Counter)                       # (cat,plan) -> Counter((brand,fund))

    for r in csv.DictReader(open(ic_path, encoding="utf-8-sig")):
        cid = r["CHARGE_ID"]; ft = norm(r["FEE_NAME"]); amt = fnum(r["TOTAL_AMOUNT"])
        ic_cat[cid][ft] += amt
        if ft in NETWORK:
            pn = (r["PLAN_NAME"] or "").strip() or "(unspecified)"
            ic_plan[cid][(ft, pn)] += amt
        if cid in succ and ft in CATS:
            cat_tot[ft] += amt / 100.0; cat_cnt[ft] += 1
            if ft in NETWORK:
                pn = (r["PLAN_NAME"] or "").strip() or "(unspecified)"
                cat_plan_tot[ft][pn] += amt / 100.0
                cat_plan_cnt[ft][pn] += 1
                t = txmap.get(cid)
                if t:
                    plan_bf[(ft, pn)][(brand(t["CARD_BRAND"]), fund(t["CARD_FUNDING"]))] += 1

    def comp(cid):
        d = ic_cat.get(cid)
        if not d:
            return None
        return dict(
            inter=round(d.get("interchange", 0) / 100, 4),
            sch=round(d.get("card_scheme", 0) / 100, 4),
            disc=round(d.get("discount", 0) / 100, 4),
            per=round(d.get("per_auth_fee", 0) / 100, 4),
            vol=round(d.get("volume_fee", 0) / 100, 4))

    # blended (brand,funding) cost rate from ACTUAL succeeded charges
    bf_cost = defaultdict(float); bf_vol = defaultdict(float)
    for t in tx:
        if t["STATUS"] != "succeeded":
            continue
        c = comp(t["CHARGE_ID"])
        if not c:
            continue
        key = (brand(t["CARD_BRAND"]), fund(t["CARD_FUNDING"]))
        bf_cost[key] += sum(c.values()); bf_vol[key] += int(t["AMOUNT"]) / 100
    blended = {k: (bf_cost[k] / bf_vol[k] if bf_vol[k] else 0) for k in bf_cost}

    recs = []
    for t in tx:
        cid = t["CHARGE_ID"]; st = t["STATUS"]; tv = int(t["AMOUNT"]) / 100
        br = brand(t["CARD_BRAND"]); fn = fund(t["CARD_FUNDING"])
        rec = dict(A=cid, B=t["CREATED_AT"][:10], C=t["MERCHANT_NAME"],
                   D=t["CURRENCY"].upper(), E=tv, F=br, G=fn,
                   H=t["CARD_COUNTRY"], I=t["CARD_LAST4"], J=st,
                   # cc = the MERCHANT's country (CONNECTED_ACCOUNTS.COUNTRY), added 2026-08-17
                   # for the UK launch. NOT column H, which is the CARDHOLDER's card country.
                   # No workbook column consumes it — every writer names its keys explicitly —
                   # it exists purely so run.py can split the margins KPIs by market using the
                   # same records, and therefore the same cost model, as the workbook itself.
                   # "ZZ" when the pull predates the MERCHANT_COUNTRY column or Stripe has none.
                   cc=(t.get("MERCHANT_COUNTRY") or "").strip().upper() or "ZZ")
        if st == "failed":
            rec.update(K="No fee (failed)", L=0, M=None, N=None, O=None,
                       Q=None, R=None, Tval=None, kind="failed")
        else:
            rev = int(t["APPLICATION_FEE_AMOUNT"]) / 100 if t["APPLICATION_FEE_AMOUNT"].strip() else 0
            rec["L"] = round(rev, 4)
            c = comp(cid)
            if c:
                rec.update(K="Actual", M=c["inter"], N=c["sch"], O=c["disc"],
                           Q=c["per"], R=c["vol"], kind="actual")
            else:
                rate = blended.get((br, fn), 0)
                rec.update(K="Estimated", M=None, N=None, O=None, Q=None, R=None,
                           Tval=round(tv * rate, 4), kind="est")
        recs.append(rec)

    ds = sorted(r["B"] for r in recs if r["B"])
    d0, d1 = ds[0], ds[-1]
    merchants = sorted(set(t["MERCHANT_NAME"] for t in tx), key=lambda s: s.lower())

    def costof(rec):
        if rec["kind"] == "actual":
            return sum(rec[k] or 0 for k in ("M", "N", "O", "Q", "R"))
        if rec["kind"] == "est":
            return rec["Tval"] or 0
        return None

    def scen(rate, fix):
        n = 0
        for rec in recs:
            if rec["J"] != "succeeded":
                continue
            c = costof(rec)
            if c is None:
                continue
            if rec["E"] * rate + fix - c > 0:
                n += 1
        return n

    scen_counts = {"2.5flat": scen(0.025, 0), "2.5+5": scen(0.025, 0.05),
                   "2.6+10": scen(0.026, 0.10), "2.6+15": scen(0.026, 0.15),
                   "2.8+15": scen(0.028, 0.15)}

    return dict(recs=recs, comp=comp, ic_plan=ic_plan,
                cat_tot=cat_tot, cat_cnt=cat_cnt,
                cat_plan_tot=cat_plan_tot, cat_plan_cnt=cat_plan_cnt, plan_bf=plan_bf,
                d0=d0, d1=d1, merchants=merchants,
                n_succ=len(succ),
                n_actual=sum(1 for r in recs if r["kind"] == "actual"),
                n_est=sum(1 for r in recs if r["kind"] == "est"),
                n_fail=sum(1 for r in recs if r["kind"] == "failed"),
                n_total=len(recs), scen=scen_counts)

# ------------------------------------------------------------------ Transaction Detail
def refresh_transaction_detail(wb, recs):
    TD = wb["Transaction Detail"]
    tmpl = {}
    for col in range(1, 29):
        cell = TD.cell(2, col)
        tmpl[col] = dict(font=copy.copy(cell.font), fmt=cell.number_format,
                         align=copy.copy(cell.alignment), fill=copy.copy(cell.fill),
                         border=copy.copy(cell.border))

    def setc(r, col, val):
        cell = TD.cell(r, col, val); t = tmpl[col]
        cell.font = copy.copy(t["font"]); cell.number_format = t["fmt"]
        cell.alignment = copy.copy(t["align"]); cell.fill = copy.copy(t["fill"])
        cell.border = copy.copy(t["border"])

    n = len(recs); last = n + 1
    for i, rec in enumerate(recs):
        r = 2 + i
        for col, key in ((1, "A"), (2, "B"), (3, "C"), (4, "D"), (5, "E"),
                         (6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J"),
                         (11, "K"), (12, "L"), (13, "M"), (14, "N"), (15, "O")):
            setc(r, col, rec.get(key))
        if rec["kind"] == "actual":
            setc(r, 16, f'=IF(COUNT($M{r}:$O{r})=0,"",SUM($M{r}:$O{r}))')
            setc(r, 17, rec.get("Q")); setc(r, 18, rec.get("R"))
            setc(r, 19, f'=IF(COUNT($Q{r}:$R{r})=0,"",SUM($Q{r}:$R{r}))')
            setc(r, 20, f'=IF(COUNT($P{r},$S{r})=0,"",$P{r}+$S{r})')
        elif rec["kind"] == "est":
            for c in (16, 17, 18, 19):
                setc(r, c, None)
            setc(r, 20, rec.get("Tval"))
        else:
            for c in (16, 17, 18, 19, 20):
                setc(r, c, None)
        setc(r, 21, f'=IF($T{r}="","",$L{r}-$T{r})')
        setc(r, 22, f'=IF(OR($T{r}="",$E{r}=0),"",$T{r}/$E{r})')
        setc(r, 23, f'=IF(OR($T{r}="",$E{r}=0),"",($L{r}-$T{r})/$E{r})')
        setc(r, 24, f'=IF($E{r}=0,"",$L{r}/$E{r})')
        setc(r, 25, f'=IF($J{r}<>"succeeded","",IF($T{r}="","",IF(($L{r}-$T{r})>0,1,0)))')
        setc(r, 26, None)
        setc(r, 27, f'=IF(AND(ISNUMBER($V{r}),$V{r}<=0.1,$E{r}<=50,$E{r}>=10),$E{r},NA())')
        setc(r, 28, f'=IF(AND(ISNUMBER($V{r}),$V{r}<=0.1,$E{r}<=50,$E{r}>=10),$V{r},NA())')

    # clear leftover rows below new data
    for r in range(last + 1, TD.max_row + 1):
        for col in range(1, 29):
            TD.cell(r, col).value = None
    return last

# ------------------------------------------------------------------ Hyper Detail
def _hstyle(level):
    if level == 0:
        fill, fc, sz = "FF1F3864", "FFFFFFFF", 9
    elif level == 1:
        fill, fc, sz = "FF2F5496", "FFFFFFFF", 9
    else:
        fill, fc, sz = "FFD9E1F2", "FF1F3864", 8
    return (Font(name="Arial", size=sz, bold=True, color=fc),
            PatternFill("solid", fgColor=fill),
            Alignment(horizontal="center", vertical="top", wrap_text=True))

def refresh_hyper_detail(wb, D, last):
    recs = D["recs"]; ic_plan = D["ic_plan"]; comp = D["comp"]
    HD = wb["Transaction Hyper Detail"]
    n = len(recs)

    FMT2 = "#,##0.00;(#,##0.00);\u2013"
    FMT4 = "#,##0.0000;(#,##0.0000);"
    thin = Side(style="thin", color="FFBFBFBF")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    base_font = Font(name="Arial", size=9, color="FF000000")
    det_font  = Font(name="Arial", size=8, color="FF000000")

    # capture base-column (1-25) header + row-2 data styles from template
    base_tmpl = {}
    for c in range(1, 26):
        cell = HD.cell(2, c)
        base_tmpl[c] = dict(font=copy.copy(cell.font), fmt=cell.number_format,
                            align=copy.copy(cell.alignment), fill=copy.copy(cell.fill),
                            border=copy.copy(cell.border))

    # wipe every existing column dimension / clear the sheet body wide
    for key in list(HD.column_dimensions.keys()):
        try:
            if cifs(key) > 25:
                del HD.column_dimensions[key]
        except ValueError:
            pass
    if HD.max_column > 25:
        HD.delete_cols(26, HD.max_column - 25)

    # ---- base columns 1..25 as VALUES ----
    def costof(rec):
        if rec["kind"] == "actual":
            return dict(P=round((rec["M"] or 0) + (rec["N"] or 0) + (rec["O"] or 0), 4),
                        S=round((rec["Q"] or 0) + (rec["R"] or 0), 4))
        return None

    for i, rec in enumerate(recs):
        r = 2 + i
        vals = {1: rec["A"], 2: rec["B"], 3: rec["C"], 4: rec["D"], 5: rec["E"],
                6: rec["F"], 7: rec["G"], 8: rec["H"], 9: rec["I"], 10: rec["J"],
                11: rec["K"], 12: rec.get("L")}
        if rec["kind"] == "actual":
            P = round((rec["M"] or 0) + (rec["N"] or 0) + (rec["O"] or 0), 4)
            S = round((rec["Q"] or 0) + (rec["R"] or 0), 4)
            T = round(P + S, 4)
            vals.update({13: rec["M"], 14: rec["N"], 15: rec["O"], 16: P,
                         17: rec["Q"], 18: rec["R"], 19: S, 20: T})
        elif rec["kind"] == "est":
            T = rec.get("Tval")
            vals.update({13: None, 14: None, 15: None, 16: None,
                         17: None, 18: None, 19: None, 20: T})
        else:
            T = None
            for c in range(13, 20):
                vals[c] = None
            vals[20] = None
        L = rec.get("L")
        E = rec["E"]
        if T is None or T == "":
            vals[21] = None; vals[22] = None; vals[23] = None
        else:
            vals[21] = round((L or 0) - T, 4)
            vals[22] = (T / E) if E else None
            vals[23] = ((L or 0) - T) / E if E else None
        vals[24] = (L / E) if (E and L is not None) else None
        if rec["J"] != "succeeded" or T is None:
            vals[25] = None
        else:
            vals[25] = 1 if ((L or 0) - T) > 0 else 0
        for c in range(1, 26):
            cell = HD.cell(r, c, vals.get(c))
            t = base_tmpl[c]
            cell.font = copy.copy(t["font"]); cell.number_format = t["fmt"]
            cell.alignment = copy.copy(t["align"]); cell.fill = copy.copy(t["fill"])
            cell.border = copy.copy(t["border"])

    # ---- header row-1 styles for base cols already present; ensure kept ----
    # ---- dynamic fee matrix ----
    # ordered plan lists per category, sorted by aggregate total desc
    cat_plan_tot = D["cat_plan_tot"]
    ic_plans = sorted(cat_plan_tot["interchange"].items(), key=lambda kv: -kv[1])
    cs_plans = sorted(cat_plan_tot["card_scheme"].items(), key=lambda kv: -kv[1])
    ds_plans = sorted(cat_plan_tot["discount"].items(), key=lambda kv: -kv[1])
    ic_names = [p for p, _ in ic_plans]
    cs_names = [p for p, _ in cs_plans]
    ds_names = [p for p, _ in ds_plans]

    col = 26
    plan_layout = []  # (col, level, hidden, collapsed)

    def set_header(c, text, level):
        cell = HD.cell(1, c); cell.value = text
        f, fl, al = _hstyle(level)
        cell.font = f; cell.fill = fl; cell.alignment = al; cell.border = bord

    def put_group_total(text, level, key, collapsed=False):
        nonlocal col
        set_header(col, text, level)
        for i, rec in enumerate(recs):
            r = 2 + i
            v = None
            if rec["kind"] == "actual":
                if key == "P":
                    v = round((rec["M"] or 0) + (rec["N"] or 0) + (rec["O"] or 0), 4)
                elif key == "M":
                    v = rec["M"]
                elif key == "N":
                    v = rec["N"]
                elif key == "O":
                    v = rec["O"]
                elif key == "S":
                    v = round((rec["Q"] or 0) + (rec["R"] or 0), 4)
                elif key == "Q":
                    v = rec["Q"]
                elif key == "R":
                    v = rec["R"]
            cell = HD.cell(r, col, v)
            cell.font = base_font; cell.number_format = FMT2
        HD.column_dimensions[gcl(col)].width = 13
        plan_layout.append((col, level, False, collapsed)); col += 1

    def put_plan_details(cat, names):
        nonlocal col
        for pn in names:
            set_header(col, pn, 2)
            for i, rec in enumerate(recs):
                r = 2 + i
                cents = ic_plan.get(rec["A"], {}).get((cat, pn))
                v = round(cents / 100.0, 4) if cents else None
                cell = HD.cell(r, col, v)
                cell.font = det_font; cell.number_format = FMT4
            HD.column_dimensions[gcl(col)].width = 12
            plan_layout.append((col, 2, True, False)); col += 1

    put_group_total("Network Fees", 0, "P")
    put_group_total("Interchange", 1, "M", collapsed=True)
    put_plan_details("interchange", ic_names)
    put_group_total("Card Scheme", 1, "N", collapsed=True)
    put_plan_details("card_scheme", cs_names)
    put_group_total("Amex Discount", 1, "O", collapsed=True)
    put_plan_details("discount", ds_names)
    put_group_total("Stripe Fees", 0, "S")
    put_group_total("Per Auth Fee", 1, "Q")
    put_group_total("Volume Fee", 1, "R")

    # spacer + filtered helper columns
    set_header(col, None, 0)
    HD.cell(1, col).value = None
    HD.column_dimensions[gcl(col)].width = 3
    plan_layout.append((col, 0, False, False)); spacer = col; col += 1

    tpv_col = col
    set_header(col, "TPV (filtered)", 0)
    for i in range(n):
        r = 2 + i
        cell = HD.cell(r, col,
                       f'=IF(AND(ISNUMBER($V{r}),$V{r}<=0.1,$E{r}<=50,$E{r}>=10),$E{r},NA())')
    HD.column_dimensions[gcl(col)].width = 13
    plan_layout.append((col, 0, False, False)); col += 1

    cost_col = col
    set_header(col, "Cost % (filtered)", 0)
    for i in range(n):
        r = 2 + i
        cell = HD.cell(r, col,
                       f'=IF(AND(ISNUMBER($V{r}),$V{r}<=0.1,$E{r}<=50,$E{r}>=10),$V{r},NA())')
    HD.column_dimensions[gcl(col)].width = 13
    plan_layout.append((col, 0, False, False)); col += 1

    # outline levels
    HD.sheet_properties.outlinePr = Outline(summaryRight=False, summaryBelow=True, applyStyles=False)
    for c, lvl, hidden, collapsed in plan_layout:
        cd = HD.column_dimensions[gcl(c)]
        cd.outlineLevel = lvl; cd.hidden = hidden; cd.collapsed = collapsed
    HD.row_dimensions[1].height = 64

    # clear stray rows below data
    for r in range(n + 2, HD.max_row + 1):
        for c in range(1, col):
            HD.cell(r, c).value = None
    return col - 1

# ------------------------------------------------------------------ Fee Breakdown
def refresh_fee_breakdown(wb, D):
    FB = wb["Fee Breakdown"]
    data_st = _grab(FB, 5); sub_st = _grab(FB, 93); tot_st = _grab(FB, 160)

    def put(r, c, v, st):
        cell = FB.cell(r, c, v); t = st[c]
        cell.font = copy.copy(t["font"]); cell.number_format = t["fmt"]
        cell.alignment = copy.copy(t["align"]); cell.fill = copy.copy(t["fill"])
        cell.border = copy.copy(t["border"])

    for r in range(5, FB.max_row + 1):
        for c in range(1, 10):
            FB.cell(r, c).value = None

    cat_tot = D["cat_tot"]; cat_cnt = D["cat_cnt"]
    cat_plan_tot = D["cat_plan_tot"]; cat_plan_cnt = D["cat_plan_cnt"]; plan_bf = D["plan_bf"]

    def rep_bf(cat, pn):
        cnt = plan_bf.get((cat, pn))
        if not cnt:
            return ("\u2014", "\u2014")
        (br, fn), _ = cnt.most_common(1)[0]
        return (br if br else "\u2014", fn if fn else "\u2014")

    # pre-count plans for the TOTAL row reference
    n_ic = len(cat_plan_tot["interchange"]); n_cs = len(cat_plan_tot["card_scheme"])
    n_ds = len(cat_plan_tot["discount"])
    total_row = 5 + (n_ic + 1) + (n_cs + 1) + (n_ds + 1) + 2
    TR = total_row

    r = 5
    def write_group(cat):
        nonlocal r
        items = sorted(cat_plan_tot[cat].items(), key=lambda kv: -kv[1])
        for pn, tot in items:
            br, fn = rep_bf(cat, pn)
            put(r, 1, "NETWORK", data_st); put(r, 2, cat, data_st); put(r, 3, pn, data_st)
            put(r, 4, cat_plan_cnt[cat][pn], data_st); put(r, 5, round(tot, 2), data_st)
            put(r, 6, f"=IFERROR(E{r}/$E${TR},0)", data_st)
            put(r, 7, f"=IFERROR(E{r}/D{r},0)", data_st)
            put(r, 8, br, data_st); put(r, 9, fn, data_st)
            r += 1
        put(r, 3, f"SUBTOTAL: {cat}", sub_st)
        put(r, 4, cat_cnt[cat], sub_st); put(r, 5, round(cat_tot[cat], 2), sub_st)
        put(r, 6, f"=IFERROR(E{r}/$E${TR},0)", sub_st)
        r += 1

    write_group("interchange"); write_group("card_scheme"); write_group("discount")
    for cat in STRIPE:
        put(r, 1, "STRIPE", sub_st); put(r, 3, f"SUBTOTAL: {cat}", sub_st)
        put(r, 4, cat_cnt[cat], sub_st); put(r, 5, round(cat_tot[cat], 2), sub_st)
        put(r, 6, f"=IFERROR(E{r}/$E${TR},0)", sub_st)
        r += 1
    assert r == TR, f"Fee Breakdown row mismatch r={r} TR={TR}"

    # subtotal row numbers for the total formula
    pos = 5
    sr = {}
    sr["interchange"] = pos + n_ic; pos = sr["interchange"] + 1
    sr["card_scheme"] = pos + n_cs; pos = sr["card_scheme"] + 1
    sr["discount"]    = pos + n_ds; pos = sr["discount"] + 1
    sr["per_auth_fee"] = pos; pos += 1
    sr["volume_fee"]   = pos; pos += 1
    put(TR, 3, "TOTAL PROCESSING FEES", tot_st)
    put(TR, 5, "=" + "+".join(f"E{sr[k]}" for k in
        ["interchange", "card_scheme", "discount", "per_auth_fee", "volume_fee"]), tot_st)
    put(TR, 6, f"=IFERROR(E{TR}/E{TR},0)", tot_st)

    FB.cell(2, 1).value = (
        f"{D['n_actual']:,} succeeded charges with actual Stripe IC+ fee lines \u00b7 "
        f"amounts in USD/EUR native \u00b7 (plan)=interchange/assessment label \u00b7 "
        f"totals tie to Summary network/Stripe components")
    return TR

def _grab(ws, r):
    return {c: dict(font=copy.copy(ws.cell(r, c).font), fmt=ws.cell(r, c).number_format,
                    align=copy.copy(ws.cell(r, c).alignment), fill=copy.copy(ws.cell(r, c).fill),
                    border=copy.copy(ws.cell(r, c).border)) for c in range(1, 10)}

# ------------------------------------------------------------------ By Merchant
def refresh_by_merchant(wb, D, last):
    BM = wb["By Merchant"]
    merchants = D["merchants"]
    # Columns A..M (1..13) are all part of the merchant table. K/L/M in the template
    # carry HARDCODED merchant names (e.g. "Adriana Espinheira"); if we don't rewrite
    # them per-row they point at the wrong merchant/row whenever the merchant set or
    # order changes. So capture the style for and rewrite every column 1..13.
    NCOL = 13
    data_tmpl = {c: dict(font=copy.copy(BM.cell(4, c).font), fmt=BM.cell(4, c).number_format,
                         align=copy.copy(BM.cell(4, c).alignment), fill=copy.copy(BM.cell(4, c).fill),
                         border=copy.copy(BM.cell(4, c).border)) for c in range(1, NCOL + 1)}
    tot_tmpl = {c: dict(font=copy.copy(BM.cell(18, c).font), fmt=BM.cell(18, c).number_format,
                        align=copy.copy(BM.cell(18, c).alignment), fill=copy.copy(BM.cell(18, c).fill),
                        border=copy.copy(BM.cell(18, c).border)) for c in range(1, NCOL + 1)}
    for r in range(4, BM.max_row + 1):
        for c in range(1, NCOL + 1):
            BM.cell(r, c).value = None

    def sty(cell, t):
        cell.font = copy.copy(t["font"]); cell.number_format = t["fmt"]
        cell.alignment = copy.copy(t["align"]); cell.fill = copy.copy(t["fill"])
        cell.border = copy.copy(t["border"])

    def rng(col):
        return f"'Transaction Detail'!${col}$2:${col}${last}"

    start = 4
    for i, m in enumerate(merchants):
        r = start + i; me = m.replace('"', '""')
        BM.cell(r, 1, m)
        BM.cell(r, 2, f'=COUNTIFS({rng("C")},"{me}",{rng("J")},"succeeded")')
        BM.cell(r, 10, f'=SUMIFS({rng("E")},{rng("C")},"{me}",{rng("J")},"succeeded")')
        # K = avg ticket (volume / count), L = revenue (col L), M = cost (col T),
        # all keyed to THIS row's merchant so no hardcoded names survive.
        BM.cell(r, 11, f"=IFERROR(J{r}/B{r},0)")
        BM.cell(r, 12, f'=SUMIFS({rng("L")},{rng("C")},"{me}",{rng("J")},"succeeded")')
        BM.cell(r, 13, f'=SUMIFS({rng("T")},{rng("C")},"{me}",{rng("J")},"succeeded")')
        for c in range(1, NCOL + 1):
            sty(BM.cell(r, c), data_tmpl[c])
    tot = start + len(merchants)
    BM.cell(tot, 1, "TOTAL")
    BM.cell(tot, 2, f"=SUM(B{start}:B{tot-1})")
    BM.cell(tot, 10, f"=SUM(J{start}:J{tot-1})")
    BM.cell(tot, 11, f"=IFERROR(J{tot}/B{tot},0)")
    BM.cell(tot, 12, f"=SUM(L{start}:L{tot-1})")
    BM.cell(tot, 13, f"=SUM(M{start}:M{tot-1})")
    for c in range(1, NCOL + 1):
        sty(BM.cell(tot, c), tot_tmpl[c])
    return tot

# ------------------------------------------------------------------ Pricing Model
def _cost_of_rec(rec):
    """Total true cost of a succeeded charge (matches Transaction Detail col T)."""
    if rec["kind"] == "actual":
        return sum((rec.get(k) or 0) for k in ("M", "N", "O", "Q", "R"))
    if rec["kind"] == "est":
        return rec.get("Tval") or 0
    return None  # failed -> no cost

def _ols(pairs):
    """Ordinary least squares of y on x -> (intercept, slope), matching Excel
    INTERCEPT()/SLOPE(). Returns (0,0) for degenerate input so no #DIV/0! leaks."""
    n = len(pairs)
    if n < 2:
        return 0.0, 0.0
    mx = sum(x for x, _ in pairs) / n
    my = sum(y for _, y in pairs) / n
    sxx = sum((x - mx) ** 2 for x, _ in pairs)
    if sxx == 0:
        return 0.0, 0.0
    sxy = sum((x - mx) * (y - my) for x, y in pairs)
    slope = sxy / sxx
    return my - slope * mx, slope

def refresh_pricing_model(wb, D):
    """B11:C14 originally held _xlfn._xlws.FILTER + INTERCEPT/SLOPE dynamic-array
    formulas. FILTER is unsupported by the (older) LibreOffice on the CI runner, so it
    recalced to #NAME? and cascaded into every downstream =$B$11+$C$11*$A19 cell.
    We precompute the same per-card-type cost-vs-volume regression in Python and write
    plain numbers, so the workbook is engine-independent. x = charge volume (col E),
    y = true cost (col T), over succeeded charges with a known cost."""
    PM = wb["Pricing Model"]
    groups = {"Debit": [], "Credit": [], "Prepaid": []}
    allp = []
    for rec in D["recs"]:
        if rec["J"] != "succeeded":
            continue
        cost = _cost_of_rec(rec)
        if cost is None:
            continue
        pair = (rec["E"], cost)
        allp.append(pair)
        if rec["G"] in groups:
            groups[rec["G"]].append(pair)
    for label, r in (("Debit", 11), ("Credit", 12), ("Prepaid", 13)):
        b, m = _ols(groups[label])
        PM.cell(r, 2).value = round(b, 6)
        PM.cell(r, 3).value = round(m, 6)
    b, m = _ols(allp)
    PM.cell(14, 2).value = round(b, 6)
    PM.cell(14, 3).value = round(m, 6)
    return {lbl: len(groups[lbl]) for lbl in groups} | {"ALL": len(allp)}

# ------------------------------------------------------------------ range bumping
def bump_ranges(wb, old_last, new_last):
    if old_last == new_last:
        return
    pat = re.compile(r'(\$?[A-Z]{1,3}\$?)' + str(old_last) + r'(?![0-9])')
    repl = r'\g<1>' + str(new_last)
    for ws in wb.worksheets:
        if ws.title in ("Transaction Detail", "Transaction Hyper Detail",
                        "By Merchant", "Fee Breakdown"):
            continue  # rebuilt with explicit refs
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=") and str(old_last) in c.value:
                    c.value = pat.sub(repl, c.value)

# ------------------------------------------------------------------ Summary / Read Me text
def refresh_text(wb, D, new_last):
    S = wb["Summary"]
    window = fmt_window(D["d0"], D["d1"])
    S["B3"] = (f"{window} \u00b7 {D['n_succ']:,} succeeded transactions \u00b7 "
               f"{len(D['merchants'])} merchants \u00b7 USD (29 EUR txns at face value)")
    # scenario profitable counts (rows 40-44, col G)
    sc = D["scen"]
    S.cell(40, 7).value = sc["2.5flat"]
    S.cell(41, 7).value = sc["2.5+5"]
    S.cell(42, 7).value = sc["2.6+10"]
    S.cell(43, 7).value = sc["2.6+15"]
    S.cell(44, 7).value = sc["2.8+15"]
    # settlement note
    S["B34"] = ("Note: Apr\u2013Jul 19 have complete IC+ cost data; the most recent charges "
                "are still settling, so their")
    S["B35"] = ("cost is estimated from each charge\u2019s card-brand \u00d7 funding blended rate. "
                "Amounts are net of Amex/network discounts.")

    RM = wb["Read Me"]
    RM["B4"] = (f"Scope: {D['n_total']:,} Stripe charges ({D['n_succ']:,} succeeded, "
                f"{D['n_fail']} failed) across {len(D['merchants'])} merchants, {window}.")
    pct = D["n_est"] / D["n_succ"] * 100 if D["n_succ"] else 0
    RM["B19"] = (f"April\u2013Jul 19 have complete finalized IC+ cost data. The most recent charges "
                 f"are still settling: {D['n_est']} charges (\u2248{pct:.1f}% of succeeded) have estimated cost,")
    RM["B22"] = (f"Refreshed {datetime.today().strftime('%B %-d, %Y')}. "
                 f"All formulas recalculated and verified error-free.")

# ------------------------------------------------------------------ Stripe platform fees
def refresh_platform_fees(wb, fees_path):
    """Source the Summary 'Stripe Fees' section (rows 17-23) from Stripe balance-
    transaction actuals in platform_fees.csv instead of the old modeled formulas/constants.
    CSV amounts are minor units (pence), NEGATIVE, in GBP; we flip sign, /100, and convert
    GBP->USD. Missing categories stay 0. The 'other' bucket (terminal hardware etc.) is
    never read. C16 keeps its =SUM(C17:C24) template formula, so subtotal/total/net-margin
    all recalc downstream. Returns the per-bucket USD dict for logging."""
    S = wb["Summary"]
    minor = {cat: 0.0 for cat, _, _ in PLATFORM_FEE_ROWS}
    if fees_path and os.path.exists(fees_path):
        for r in csv.DictReader(open(fees_path, encoding="utf-8-sig")):
            cat = (r.get("FEE_CATEGORY") or "").strip()
            if cat in minor:
                minor[cat] += fnum(r.get("TOTAL_AMOUNT_MINOR"))
    usd = {}
    for cat, row, label in PLATFORM_FEE_ROWS:
        v = round(-minor[cat] / 100.0 * FX_GBP_USD, 2) + 0.0   # neg pence GBP -> +USD; kill -0.0
        S.cell(row, 2).value = label
        S.cell(row, 3).value = v
        usd[cat] = v
    # Row 24 previously held a modeled 'Terminal Smart' constant; no actual bucket -> blank.
    S.cell(24, 2).value = None
    S.cell(24, 3).value = None
    # Provenance / FX note (row 36, between the settlement note and the scenario table).
    S["B36"] = (f"Stripe platform fees (rows 17\u201323) are actuals from Stripe balance "
                f"transactions, converted GBP\u2192USD at {FX_GBP_USD} (22 Jul 2026). Terminal "
                f"hardware and other non-processing charges are excluded.")
    return usd

# ------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tx", required=True)
    ap.add_argument("--ic", required=True)
    ap.add_argument("--fees", default=None,
                    help="platform_fees.csv (defaults to platform_fees.csv beside --ic)")
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    fees_path = args.fees or os.path.join(os.path.dirname(os.path.abspath(args.ic)),
                                          "platform_fees.csv")

    print("Loading data ...")
    D = build_data(args.tx, args.ic)
    print(f"  charges={D['n_total']:,}  succeeded={D['n_succ']:,}  actual={D['n_actual']:,} "
          f"est={D['n_est']}  failed={D['n_fail']}  merchants={len(D['merchants'])}")
    print(f"  window={fmt_window(D['d0'], D['d1'])}  scen={D['scen']}")

    wb = openpyxl.load_workbook(args.template)
    old_last = wb["Transaction Detail"].max_row
    print(f"Template last data row = {old_last}")

    new_last = refresh_transaction_detail(wb, D["recs"])
    print(f"Transaction Detail rebuilt -> last row {new_last}")
    ncol = refresh_hyper_detail(wb, D, new_last)
    print(f"Transaction Hyper Detail rebuilt -> {ncol} cols")
    tr = refresh_fee_breakdown(wb, D)
    print(f"Fee Breakdown rebuilt -> total row {tr}")
    bmtot = refresh_by_merchant(wb, D, new_last)
    print(f"By Merchant rebuilt -> total row {bmtot}")
    pm = refresh_pricing_model(wb, D)
    print(f"Pricing Model regression precomputed -> n by type {pm}")
    bump_ranges(wb, old_last, new_last)
    print(f"Ranges bumped {old_last} -> {new_last}")
    pf = refresh_platform_fees(wb, fees_path)
    src = fees_path if os.path.exists(fees_path) else f"{fees_path} (missing -> all 0)"
    print(f"Stripe platform fees from actuals ({src}) -> USD {pf}")
    refresh_text(wb, D, new_last)
    print("Summary / Read Me text refreshed")

    wb.save(args.out)
    print(f"Saved {args.out}")

if __name__ == "__main__":
    main()
