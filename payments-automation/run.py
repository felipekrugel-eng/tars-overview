#!/usr/bin/env python3
"""
FACADASH-pattern server-side orchestrator for the Embedded Payments true-cost workbook.

Single entrypoint the GitHub Action calls (mirrors kpi-automation/run.py). Does the
WHOLE refresh on the runner — no desktop, no manual steps:

  1. PULL     Snowflake -> data/transactions.csv + data/icplus_costs.csv
              (pull_payments.py; sanity-gated). Skippable with --no-pull for local
              testing against the committed seed CSVs.
  2. REBUILD  refresh_workbook.py rebuilds every sheet of the V2 workbook from the CSVs.
  3. RECALC   LibreOffice headless recalculates so Excel opens with live values.
  4. VERIFY   scan for GENUINE formula errors (excluding the intentional #N/A NA()
              sentinels in the filtered TPV/Cost% helper columns) and confirm all 13
              sheets + headline cells are populated. Any failure -> non-zero exit so
              the workflow does NOT commit a broken workbook.

On success the finished workbook is written to output/loyverse_payments_analysis_ful_V2.xlsx,
which the workflow commits to the repo. That committed file is the source of truth,
refreshed daily with zero dependence on anyone's computer.
"""
import os, sys, subprocess, shutil, pathlib, argparse, json, datetime
import openpyxl

HERE     = pathlib.Path(__file__).resolve().parent
DATA_DIR = pathlib.Path(os.environ.get("DATA_DIR", HERE / "data"))
TEMPLATE = pathlib.Path(os.environ.get("TEMPLATE",
                        HERE / "template" / "loyverse_payments_analysis_ful_V2.xlsx"))
OUT_DIR  = pathlib.Path(os.environ.get("OUT_DIR", HERE / "output"))
OUT_XLSX = OUT_DIR / "loyverse_payments_analysis_ful_V2.xlsx"
# The "Summary of Margins" dashboard page (activated-payments/index.html) reads this file.
# Committed by the workflow alongside the workbook; Netlify auto-deploys on push.
MARGINS_OUT = pathlib.Path(os.environ.get("MARGINS_OUT",
                           HERE.parent / "activated-payments" / "margins-data.js"))
BUILD_XLSX = HERE / "_build.xlsx"   # scratch (pre-recalc)

GENUINE_ERRS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"}
EXPECTED_SHEETS = 13
HEADLINE = ["C5", "C6", "C8", "C11", "C12", "C25"]


def sh(cmd, **kw):
    print("+", " ".join(str(c) for c in cmd), flush=True)
    subprocess.run(cmd, check=True, **kw)


def soffice_bin():
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    for p in ("/usr/bin/soffice", "/usr/bin/libreoffice"):
        if os.path.exists(p):
            return p
    raise RuntimeError("LibreOffice (soffice) not found on PATH.")


def recalc(src, out_dir):
    """LibreOffice headless resave -> forces a full recalc and caches values."""
    out_dir.mkdir(parents=True, exist_ok=True)
    sh([soffice_bin(), "--headless", "--calc", "--convert-to", "xlsx",
        "--outdir", str(out_dir), str(src)])
    produced = out_dir / (src.stem + ".xlsx")
    if not produced.exists():
        raise RuntimeError(f"LibreOffice did not produce {produced}")
    return produced


def verify(path):
    # data_only=True gives the LibreOffice-cached VALUES; a second load with
    # data_only=False lets us print the offending FORMULA next to each error cell,
    # which is what actually pinpoints the root cause on the runner.
    wb = openpyxl.load_workbook(path, data_only=True)
    fwb = openpyxl.load_workbook(path, data_only=False)
    sheets = [ws.title for ws in wb.worksheets]
    genuine, na = 0, 0
    err_cells = []   # [(sheet, coord, err_value, formula)]
    for ws in wb.worksheets:
        fws = fwb[ws.title]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str):
                    if v == "#N/A":
                        na += 1
                    elif v in GENUINE_ERRS:
                        genuine += 1
                        formula = fws[c.coordinate].value
                        err_cells.append((ws.title, c.coordinate, v, formula))
    problems = []
    if len(sheets) != EXPECTED_SHEETS:
        problems.append(f"sheet count {len(sheets)} != {EXPECTED_SHEETS} ({sheets})")
    if genuine:
        problems.append(f"{genuine} genuine formula errors")
        # Group by sheet+error-type so the log is scannable, then dump the first
        # ~40 concrete cells with their formulas.
        from collections import Counter
        by_sheet = Counter((s, e) for s, _, e, _ in err_cells)
        print("VERIFY: genuine error breakdown (sheet, type -> count):", flush=True)
        for (sheet, etype), n in sorted(by_sheet.items(), key=lambda x: -x[1]):
            print(f"    {sheet!r:28} {etype:8} x{n}", flush=True)
        print("VERIFY: first offending cells (sheet!cell = err  <- formula):", flush=True)
        for sheet, coord, ev, formula in err_cells[:40]:
            print(f"    {sheet}!{coord} = {ev}   <- {formula}", flush=True)
        if len(err_cells) > 40:
            print(f"    ... +{len(err_cells)-40} more", flush=True)
    s = wb["Summary"]
    blanks = [a for a in HEADLINE if s[a].value in (None, "")]
    if blanks:
        problems.append(f"blank headline cells: {blanks}")
    headline = {a: s[a].value for a in HEADLINE}
    print(f"VERIFY: sheets={len(sheets)} genuine_errors={genuine} "
          f"intentional_na={na} headline={headline}", flush=True)
    return problems


def _num(v):
    """Cell value -> float or None ('-' placeholder / blank -> None)."""
    if v in (None, "", "-", "\u2013"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def margins_by_country(tx_path, ic_path, stripe_fees_total):
    """Split the Summary KPIs by MERCHANT country for the dashboard's country filter.

    Deliberately NOT a per-country rebuild of the 13-sheet workbook: the workbook stays the
    single source of truth and its Summary is untouched, so the all-countries view can never
    restate. This recomputes the same KPIs from the same records that built the workbook — it
    imports refresh_workbook.build_data rather than reimplementing the cost model, so a country
    slice cannot drift from the total by using a different definition of "cost".

    WHAT SPLITS EXACTLY, AND WHAT DOES NOT:
      * txns, tpv, revenue, network fees (interchange / card scheme / Amex discount) and the
        profitability counts are per-charge, join to the merchant, and split exactly. Summing
        them across countries reproduces the workbook Summary to the cent.
      * Stripe's own platform fees (per-auth, volume, Radar, Tap to Pay, payout, terminal use)
        come from query3, which reads the PLATFORM account's balance transactions. Those rows
        carry no connected-account id at all, so they genuinely cannot be attributed to a
        merchant, let alone a country. They are apportioned pro-rata by each country's share of
        TPV and flagged as such (stripeFeesBasis). Any figure downstream of them — totalFees,
        netMargin, netTakeRate — is therefore part-actual, part-apportioned PER COUNTRY, and is
        exact only for the all-countries total. The dashboard must say so.
    """
    sys.path.insert(0, str(HERE))
    from refresh_workbook import build_data          # same cost model as the workbook
    D = build_data(tx_path, ic_path)

    tpv_total = sum(r["E"] for r in D["recs"] if r["J"] == "succeeded")
    by = {}
    def slot(cc):
        return by.setdefault(cc, dict(
            txns=0, tpv=0.0, revenue=0.0, interchange=0.0, cardScheme=0.0, amexDiscount=0.0,
            profitableTxns=0, withActual=0, withEstimated=0, failed=0))

    for r in D["recs"]:
        o = slot(r.get("cc") or "ZZ")
        if r["kind"] == "failed":
            o["failed"] += 1
            continue
        if r["kind"] == "actual":
            o["withActual"] += 1
        elif r["kind"] == "est":
            o["withEstimated"] += 1
        if r["J"] != "succeeded":
            continue
        o["txns"] += 1
        o["tpv"] += r["E"]
        o["revenue"] += r["L"] or 0
        o["interchange"]  += r.get("M") or 0
        o["cardScheme"]   += r.get("N") or 0
        o["amexDiscount"] += r.get("O") or 0
        # Same test as Transaction Detail col Y: revenue minus per-charge true cost > 0.
        # The margin is rounded to 6dp before the comparison because the fee components are
        # stored to 4dp and binary floats do not sum back exactly — 0.27 comes out as
        # 0.26999999999999996, which would score two exactly-break-even charges as profitable
        # and put this count 2 above the workbook's.
        cost = (sum(r.get(k) or 0 for k in ("M", "N", "O", "Q", "R"))
                if r["kind"] == "actual" else r.get("Tval"))
        if cost is not None and round((r["L"] or 0) - cost, 6) > 0:
            o["profitableTxns"] += 1

    out = {}
    for cc, o in sorted(by.items()):
        network = o["interchange"] + o["cardScheme"] + o["amexDiscount"]
        share = (o["tpv"] / tpv_total) if tpv_total else 0.0
        stripe = (stripe_fees_total or 0.0) * share
        total_fees = network + stripe
        net = o["revenue"] - total_fees
        out[cc] = {
            "kpis": {
                "txns": o["txns"],
                "tpv": round(o["tpv"], 2),
                "avgTicket": (o["tpv"] / o["txns"]) if o["txns"] else None,
                "revenue": round(o["revenue"], 2),
                "takeRate": (o["revenue"] / o["tpv"]) if o["tpv"] else None,
                "totalFees": round(total_fees, 4),
                "netMargin": round(net, 4),
                "netTakeRate": (net / o["tpv"]) if o["tpv"] else None,
                "pctProfitable": (o["profitableTxns"] / o["txns"]) if o["txns"] else None,
                "profitableTxns": o["profitableTxns"],
                "unprofitableTxns": o["txns"] - o["profitableTxns"],
                "withActual": o["withActual"],
                "withEstimated": o["withEstimated"],
                "failed": o["failed"],
            },
            "fees": {
                "total": round(total_fees, 4),
                "network": {"total": round(network, 4),
                            "items": [round(o["interchange"], 4),
                                      round(o["cardScheme"], 4),
                                      round(o["amexDiscount"], 4)]},
                "stripe": {"total": round(stripe, 4)},
            },
            # Scenario revenue is a pure function of this country's own volume and count,
            # exactly as the Summary's formulas are of C6/C5 — so these are exact.
            "scenarios": [round(x, 5) for x in (
                o["revenue"],
                o["tpv"] * 0.025,
                o["tpv"] * 0.025 + o["txns"] * 0.05,
                o["tpv"] * 0.026 + o["txns"] * 0.10,
                o["tpv"] * 0.026 + o["txns"] * 0.15,
                o["tpv"] * 0.028 + o["txns"] * 0.15)],
            "tpvShare": round(share, 6),
            "stripeFeesBasis": "apportioned_by_tpv",
        }
    return out


def emit_margins(xlsx_path, out_path, tx_path=None, ic_path=None):
    """Serialize the workbook's Summary sheet into margins-data.js for the dashboard's
    'Summary of Margins' page. The dashboard mirrors the Summary tab exactly, so we read
    label (col B) + value (col C) straight from the deterministic Summary layout."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    s = wb["Summary"]
    B = lambda r: s.cell(r, 2).value          # label column
    C = lambda r: s.cell(r, 3).value          # value column
    pair = lambda r: {"label": B(r), "value": _num(C(r))}

    data = {
        "updated": datetime.date.today().isoformat(),
        "title": B(2),
        "subtitle": B(3),
        "kpis": {
            "txns":        _num(C(5)),
            "tpv":         _num(C(6)),
            "avgTicket":   _num(C(7)),
            "revenue":     _num(C(8)),
            "takeRate":    _num(C(9)),
            "totalFees":   _num(C(11)),
            "netMargin":   _num(C(25)),
            "netTakeRate": _num(C(26)),
            "pctProfitable":    _num(C(29)),
            "profitableTxns":   _num(C(27)),
            "unprofitableTxns": _num(C(28)),
            "withActual":    _num(C(30)),
            "withEstimated": _num(C(31)),
            "failed":        _num(C(32)),
        },
        "fees": {
            "total": _num(C(11)),
            "network": {"label": B(12), "total": _num(C(12)),
                        "items": [pair(13), pair(14), pair(15)]},
            "stripe":  {"label": B(16), "total": _num(C(16)),
                        "items": [pair(17), pair(18)]},
        },
        # Additional Stripe platform fees from actuals (rows 19-23): Radar, Tap to Pay,
        # payout, terminal-use, account-volume. Merged under "Stripe fees" in the P&L.
        # Zero-value buckets render as $0.00 ("0 until it appears"); row 24 is unused.
        "extraFees": [pair(r) for r in range(19, 24)],
        "scenarios": {
            "headers": {"name": B(38), "revenue": C(38)},
            "rows": [{"name": B(r), "revenue": _num(C(r))} for r in range(39, 45)],
        },
        "notes": {
            "settlement": " ".join(x for x in (B(34), B(35)) if x),
            "scenario": B(45),
        },
    }

    # COUNTRY FILTER (added 2026-08-17, UK launch). Everything above is the workbook Summary
    # and stays all-countries and untouched; `byCountry` is purely additive, so a reader that
    # ignores it sees precisely the numbers it saw before. Guarded: the country split must never
    # be able to stop margins-data.js being written, since the page worked without it.
    data["byCountry"] = {}
    data["countrySplitNote"] = None
    if tx_path and ic_path:
        try:
            data["byCountry"] = margins_by_country(tx_path, ic_path, _num(C(16)) or 0.0)
            data["countrySplitNote"] = (
                "Transactions, volume, revenue and network fees are actuals for the selected "
                "country. Stripe's platform fees are billed to the Loyverse platform account "
                "with no merchant attribution, so they are apportioned by share of volume — "
                "net margin for a single country is therefore an estimate. The all-countries "
                "view is exact.")
            ccs = {cc: v["kpis"]["txns"] for cc, v in data["byCountry"].items()}
            print(f"   country split: {ccs}", flush=True)
        except Exception as e:
            print(f"   !! country split skipped (non-fatal): {e}", file=sys.stderr, flush=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    js = ("// AUTO-GENERATED by payments-automation/run.py — do not edit by hand.\n"
          "// Mirrors the Summary tab of loyverse_payments_analysis_ful_V2.xlsx.\n"
          "window.MARGINS = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n")
    out_path.write_text(js, encoding="utf-8")
    print(f"OK -> {out_path}  (netMargin={data['kpis']['netMargin']}, "
          f"scenarios={len(data['scenarios']['rows'])})", flush=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-pull", action="store_true",
                    help="skip the Snowflake pull; rebuild from the committed CSVs "
                         "(local testing only)")
    args = ap.parse_args()

    tx = DATA_DIR / "transactions.csv"
    ic = DATA_DIR / "icplus_costs.csv"
    fees = DATA_DIR / "platform_fees.csv"   # Stripe platform-fee actuals (may be absent)

    # 1. PULL
    if args.no_pull:
        print("--no-pull: using committed CSVs as-is.")
    else:
        try:
            sh([sys.executable, str(HERE / "pull_payments.py")])
        except subprocess.CalledProcessError as e:
            if e.returncode == 2:
                print("Pull hit the sanity gate (grant likely pending) — "
                      "rebuilding from the last-good committed CSVs.", flush=True)
            else:
                raise

    if not tx.exists() or not ic.exists():
        print(f"ERROR: missing input CSVs ({tx}, {ic}).", file=sys.stderr)
        sys.exit(1)

    # 2. REBUILD
    sh([sys.executable, str(HERE / "refresh_workbook.py"),
        "--tx", str(tx), "--ic", str(ic), "--fees", str(fees),
        "--template", str(TEMPLATE), "--out", str(BUILD_XLSX)])

    # 3. RECALC
    recalced = recalc(BUILD_XLSX, HERE / "_recalc")

    # 4. VERIFY
    problems = verify(recalced)
    if problems:
        print("VERIFY FAILED: " + "; ".join(problems), file=sys.stderr)
        sys.exit(1)

    # publish
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(recalced, OUT_XLSX)
    print(f"OK -> {OUT_XLSX}")

    # emit the dashboard's "Summary of Margins" data file from the finished workbook.
    # The CSVs go along for the ride so the same records can be split by merchant country.
    emit_margins(OUT_XLSX, MARGINS_OUT, tx, ic)

    # tidy scratch (leave nothing to accidentally commit)
    for p in (BUILD_XLSX,):
        try: p.unlink()
        except FileNotFoundError: pass
    shutil.rmtree(HERE / "_recalc", ignore_errors=True)


if __name__ == "__main__":
    main()
