#!/usr/bin/env python3
"""
run.py — margin-automation pipeline: rebuild -> recalc -> verify -> publish.

Modelled on payments-automation/run.py and deliberately reuses its CSVs, so this
adds no Snowflake cost: payments-pull already exports transactions.csv and
icplus_costs.csv every three hours.

    REBUILD   refresh_margin_workbook.py writes the data sheets into the template
    RECALC    LibreOffice headless, so the file opens with live values
    VERIFY    the gates below. If any fails, nothing is published and the
              half-built file survives for inspection.
    PUBLISH   copy over output/<workbook>.xlsx

Env (all optional, sensible defaults):
    TX, IC          input CSVs
    TEMPLATE        template workbook
    OUT_DIR         where the finished workbook lands
    SOFFICE         path to soffice/libreoffice
"""

import os
import pathlib
import shutil
import subprocess
import sys

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent
NAME = "Loyverse_Payments_Transaction_Margin_vs_Competitors.xlsx"

TX = pathlib.Path(os.environ.get("TX", HERE / "data" / "transactions.csv"))
IC = pathlib.Path(os.environ.get("IC", HERE / "data" / "icplus_costs.csv"))
TEMPLATE = pathlib.Path(os.environ.get("TEMPLATE", HERE / "template" / NAME))
OUT_DIR = pathlib.Path(os.environ.get("OUT_DIR", HERE / "output"))
OUT_XLSX = OUT_DIR / NAME

BUILD = HERE / "_build.xlsx"
RECALC_DIR = HERE / "_recalc"

DETAIL = "Transaction Detail"
PROTECTED = ("Rate Card",)

# Rate Card input cells that must be byte-identical to the template after a run.
# Rows 6-15 are the competitor/hypothetical rates (including Felipe's 24 Aug
# corrections); rows 27-35 are the Stripe IC+ tier table transcribed from the signed
# pricing agreement; rows 40-41 are the tiered-by-card-type scenarios.
RATE_CARD_CELLS = ([f"{c}{r}" for r in range(6, 16) for c in "CD"]
                   + [f"{c}{r}" for r in range(27, 36) for c in "BCDEFG"]
                   + [f"{c}{r}" for r in (40, 41) for c in "BCDE"])


def sh(*cmd, **kw):
    print("+ " + " ".join(str(c) for c in cmd), flush=True)
    subprocess.run([str(c) for c in cmd], check=True, **kw)


def publish(src, dst):
    """Replace dst with src atomically.

    On the CI runner nothing has the file open and a plain copy would do. This matters
    for a LOCAL MIRROR: shutil.copyfile opens the destination for writing and TRUNCATES
    it, so a reader that had it open sees a half-written file for the duration. Writing
    a sibling temp file and then os.replace() makes the swap a single atomic rename -
    a reader either gets the whole old file or the whole new one, never a torn one.

    An Excel lock file (~$name.xlsx) means someone has the workbook open. Publishing
    anyway is safe for the file on disk, but Excel still holds the OLD content in
    memory and saving from there would overwrite the fresh build with stale data. So
    skip, and say so, rather than set up a silent lost update.
    """
    dst = pathlib.Path(dst)
    lock = dst.with_name("~$" + dst.name)
    if lock.exists():
        print(f"SKIPPING PUBLISH: {dst.name} is open in Excel ({lock.name} present).\n"
              f"  The build is good and is in {src}. Close the workbook and re-run.\n"
              "  Publishing now would leave Excel holding stale content that a save "
              "would write back over this build.", file=sys.stderr)
        return False
    tmp = dst.with_name(dst.name + ".tmp")
    shutil.copyfile(src, tmp)
    os.replace(tmp, dst)
    return True


def soffice():
    for n in (os.environ.get("SOFFICE"), "soffice", "libreoffice"):
        if n and shutil.which(n):
            return shutil.which(n)
    sys.exit("no LibreOffice on PATH — needed to recalculate the workbook")


def recalc(src):
    if RECALC_DIR.exists():
        shutil.rmtree(RECALC_DIR)
    RECALC_DIR.mkdir(parents=True)
    sh(soffice(), "--headless", "--norestore", "--convert-to", "xlsx",
       "--outdir", RECALC_DIR, src)
    out = RECALC_DIR / src.name
    if not out.exists():
        sys.exit(f"recalc produced nothing at {out}")
    return out


def verify(path, template):
    """Return a list of problems. Empty list means publish."""
    problems = []
    wb = openpyxl.load_workbook(path, data_only=True)
    tpl = openpyxl.load_workbook(template, data_only=False)

    # ---- 1. no formula errors anywhere
    errs = {}
    for ws in wb.worksheets:
        n = sum(1 for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("#"))
        if n:
            errs[ws.title] = n
    if errs:
        problems.append(f"formula errors: {errs}")

    # ---- 2. every sheet still present
    missing = set(tpl.sheetnames) - set(wb.sheetnames)
    if missing:
        problems.append(f"sheets missing: {sorted(missing)}")

    # ---- 3. header row 2 of Transaction Detail matches the template cell for cell.
    # USD Card Mix and Merchant Analysis address this sheet BY LETTER, so a
    # one-column drift would silently re-point every SUMIFS instead of erroring.
    d, dt = wb[DETAIL], tpl[DETAIL]
    got = [d.cell(2, c).value for c in range(1, dt.max_column + 1)]
    want = [dt.cell(2, c).value for c in range(1, dt.max_column + 1)]
    if got != want:
        bad = [f"col {c + 1}: {got[c]!r} != {want[c]!r}"
               for c in range(len(want)) if got[c] != want[c]][:5]
        problems.append("Transaction Detail header drift: " + "; ".join(bad))

    # ---- 4. Rate Card untouched
    for s in PROTECTED:
        for ref in RATE_CARD_CELLS:
            a, b = wb[s][ref].value, tpl[s][ref].value
            if a != b:
                problems.append(f"{s}!{ref} changed: {a!r} != {b!r} — a run must never "
                                "alter hand-entered rates")
                break

    # ---- 5. rows present
    last = max((r for r in range(3, d.max_row + 1) if d.cell(r, 1).value), default=0)
    nrows = max(last - 2, 0)
    if nrows < 100:
        problems.append(f"only {nrows} data rows — refusing to publish a near-empty workbook")

    # ---- 6. the identities, and the units check that actually catches scale bugs.
    # margin == revenue - cost holds at ANY scale, so it cannot detect a cents/dollars
    # error. cost as a share of volume can: it must land near 2.5%.
    g = lambda col: sum(d.cell(r, col).value or 0 for r in range(3, last + 1)
                        if isinstance(d.cell(r, col).value, (int, float)))
    vol, rev = g(7), g(14)
    ic, cs, ax, pa, vf = g(15), g(16), g(17), g(19), g(20)
    cost, marg = g(22), g(23)
    if vol <= 0:
        problems.append("total volume is zero")
    else:
        if abs(marg - (rev - cost)) > 0.05:
            problems.append(f"margin {marg:.2f} != revenue-cost {rev - cost:.2f}")
        if abs(cost - (ic + cs + ax + pa + vf)) > 0.05:
            problems.append(f"cost {cost:.2f} != sum of components {ic + cs + ax + pa + vf:.2f}")
        ratio = cost / vol
        if not 0.015 <= ratio <= 0.040:
            problems.append(f"cost is {ratio:.2%} of volume — outside 1.5%-4.0%. "
                            "Almost always an IC+ units error (TOTAL_AMOUNT is CENTS).")
        take = rev / vol
        if not 0.005 <= take <= 0.060:
            problems.append(f"take rate is {take:.2%} of volume — outside 0.5%-6.0%")

    print(f"\nVERIFY: sheets={len(wb.sheetnames)} rows={nrows:,} "
          f"volume=${vol:,.2f} revenue=${rev:,.2f} cost=${cost:,.2f} "
          f"({cost / vol:.3%} of volume) margin=${marg:,.2f}", flush=True)
    return problems


def main():
    for p in (TX, IC, TEMPLATE):
        if not p.exists():
            sys.exit(f"missing input: {p}")

    print("=== REBUILD ===", flush=True)
    sh(sys.executable, HERE / "refresh_margin_workbook.py",
       "--tx", TX, "--ic", IC, "--template", TEMPLATE, "--out", BUILD)

    print("\n=== RECALC ===", flush=True)
    recalced = recalc(BUILD)

    print("\n=== VERIFY ===", flush=True)
    problems = verify(recalced, TEMPLATE)
    if problems:
        print("VERIFY FAILED:", file=sys.stderr)
        for p in problems:
            print("  - " + p, file=sys.stderr)
        # Leave _build.xlsx and _recalc/ in place so the failing file can be opened.
        sys.exit(1)
    print("VERIFY OK", flush=True)

    print("\n=== PUBLISH ===", flush=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if not publish(recalced, OUT_XLSX):
        sys.exit(2)
    print(f"published {OUT_XLSX} ({OUT_XLSX.stat().st_size / 1e6:.1f} MB)")

    # Best-effort tidy. Never fail a good run over scratch files: some mounts
    # (e.g. a synced folder) refuse unlink even when the publish itself succeeded.
    try:
        BUILD.unlink(missing_ok=True)
    except OSError as e:
        print(f"note: could not remove {BUILD.name} ({e.strerror}) — harmless")
    shutil.rmtree(RECALC_DIR, ignore_errors=True)


if __name__ == "__main__":
    main()
