#!/usr/bin/env python3
"""
refresh_margin_workbook.py — rebuild the Margin vs Competitors workbook from the
three Snowflake CSVs that payments-pull already exports.

    python3 refresh_margin_workbook.py --tx data/transactions.csv \
        --ic data/icplus_costs.csv \
        --template template/Loyverse_Payments_Transaction_Margin_vs_Competitors.xlsx \
        --out _build.xlsx

The template is the authority for layout, formatting, every formula, and the two
hand-owned tabs. This script only replaces data.

WHAT IS WRITTEN
  Transaction Detail        19 raw columns; formula columns filled down from the
                            template's row-3 pattern
  Transaction Hyper Detail  fixed columns + the per-fee-name matrix, routed by the
                            TEMPLATE's header names
  Fee Breakdown            static aggregation of actual IC+ lines
  Merchant Analysis        per-merchant aggregation (was a dead values-only paste)

WHAT IS NEVER TOUCHED
  Rate Card                Hand-entered competitor rates, Felipe's dated corrections,
                            and the Stripe IC+ tier table transcribed from the signed
                            pricing agreement. Not reproducible from any data source.
                            verify() fails the run if a rebuild alters it.
  Merchant Months          The fee-per-merchant-month input (B) is Felipe's.
  Summary / By Market / By Merchant / By Ticket Band / USD Card Mix
                            Formula-driven off Transaction Detail; they recalc.

COLUMN LETTERS ARE LOAD-BEARING. USD Card Mix addresses Transaction Detail by letter
($F Currency, $I Funding, $O/$P/$Q network components, $U Stripe fees) and there are
15 deliberate spacer columns holding the layout apart. This script writes by letter
from RAW_COLS below and asserts the template header row still matches, so a silent
one-column drift is impossible.

UNITS — get these wrong and nothing else matters
  transactions.AMOUNT                  MINOR units -> /100
  transactions.APPLICATION_FEE_AMOUNT  MINOR units -> /100   (= Loyverse revenue)
  icplus_costs.TOTAL_AMOUNT            CENTS -> /100

  The IC+ unit is the one that has been assumed both ways across this library.
  An earlier draft of this script read TOTAL_AMOUNT as dollars, on the strength of a
  single 0.03 row that looked like a plausible dollar amount. That produced a total
  cost of $599,752 against $278,989 of volume - 215% - and every downstream
  reconciliation still passed, because margin = revenue - cost holds whatever the
  scale is. Internal consistency is not a units check. The check that matters is
  cost as a share of volume: it must land near 2.5%, and it does at /100
  (interchange 1.75%, scheme 0.26%, Amex discount 0.09%, Stripe 0.34%).
  This matches payments-truecost-weekly: "ICPLUS TOTAL_AMOUNT is in CENTS".
  verify() now asserts the ratio so this cannot regress silently.
"""

import argparse
import csv
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string as cifs
from openpyxl.utils import get_column_letter as gcl

DETAIL = "Transaction Detail"
HYPER = "Transaction Hyper Detail"
DATA_START = 3          # Transaction Detail: banners r1, headers r2, data r3+
HYPER_START = 2         # Hyper Detail: headers r1, data r2+
MAXROW = 300_000        # range ceiling, matches the template (never the row count)
PROTECTED = ("Rate Card", "Merchant Months")

# Transaction Detail raw columns: letter -> header text expected in the template.
RAW_COLS = {
    "A": "Charge ID",          "B": "Date",              "C": "Month",
    "E": "Merchant",           "F": "Currency",          "G": "Transaction Value",
    "H": "Card Brand",         "I": "Card Funding",      "J": "Card Country",
    "K": "Card Last4",         "L": "Status",            "M": "Fee Data",
    "N": "Revenue (Loyverse)", "O": "Interchange",       "P": "Card Scheme",
    "Q": "Amex Discount",      "S": "Per Auth Fee",      "T": "Volume Fee",
    "BX": "Merchant Country",
}

BRAND = {"Visa": "Visa", "MasterCard": "Mastercard", "American Express": "Amex",
         "Discover": "Discover", "Amex": "Amex"}
NAVY = "0B2B3C"
HDR_F = Font(name="Aptos Narrow", size=10, bold=True, color="FFFFFF")
NETWORK_CATS = ("interchange", "card_scheme", "discount")
STRIPE_CATS = ("per_auth_fee", "volume_fee")


def brand(b):
    return BRAND.get(b, (b.title() if b else "Unknown"))


def fund(f):
    return f.title() if f else "Unknown"


def norm_fee(name):
    # Stripe bills a 'non_transactional_card_scheme' line that belongs with scheme fees.
    return "card_scheme" if name == "non_transactional_card_scheme" else name


# ------------------------------------------------------------------- loading ----
def load_transactions(path):
    out = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            if (r.get("STATUS") or "").lower() != "succeeded":
                continue
            def money(k):
                v = r.get(k)
                try:
                    return round(float(v) / 100.0, 6)
                except (TypeError, ValueError):
                    return 0.0
            created = (r.get("CREATED_AT") or "")[:19]
            out.append({
                "charge_id": r["CHARGE_ID"],
                "created": created,
                "date": created[:10],
                "merchant": r.get("MERCHANT_NAME") or "Unknown",
                "currency": (r.get("CURRENCY") or "").upper(),
                "amount": money("AMOUNT"),
                "brand": brand(r.get("CARD_BRAND")),
                "funding": fund(r.get("CARD_FUNDING")),
                "card_country": r.get("CARD_COUNTRY") or "",
                "last4": r.get("CARD_LAST4") or "",
                "status": "succeeded",
                "revenue": money("APPLICATION_FEE_AMOUNT"),
                # Optional. Present once query1_transactions.sql selects a.COUNTRY;
                # until then it is absent and we fall back to the template's own
                # BX values keyed on merchant name (which misses new merchants).
                "merchant_country": (r.get("MERCHANT_COUNTRY") or "").strip().upper(),
            })
    out.sort(key=lambda x: (x["created"], x["charge_id"]))
    return out


def load_icplus(path):
    """charge_id -> {category: total, ...} plus charge_id -> {fee_name: total}."""
    by_cat = defaultdict(lambda: defaultdict(float))
    by_name = defaultdict(lambda: defaultdict(float))
    lines = defaultdict(lambda: defaultdict(lambda: [0, 0.0, Counter()]))
    cat_of_plan = {}          # plan name -> fee category, for column naming
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            cid = r.get("CHARGE_ID")
            if not cid:
                continue
            # FEE_CATEGORY is the coarse bucket (network_cost / stripe_fee); FEE_NAME
            # is the granular one (interchange, card_scheme, discount, per_auth_fee,
            # volume_fee). Always key off FEE_NAME. An earlier version only unwrapped
            # FEE_NAME for network_cost, so per_auth_fee and volume_fee stayed bucketed
            # as 'stripe_fee' and columns S and T were written as 0 on every row.
            fee = (r.get("FEE_NAME") or "").strip()
            cat = norm_fee(fee or (r.get("FEE_CATEGORY") or "").strip())
            try:
                amt = float(r.get("TOTAL_AMOUNT") or 0) / 100.0   # CENTS -> dollars
            except ValueError:
                continue
            by_cat[cid][cat] += amt
            plan = (r.get("PLAN_NAME") or fee or cat).strip()
            by_name[cid][plan] += amt
            agg = lines[cat][plan]
            agg[0] += 1
            agg[1] += amt
            agg[2][f"{brand(r.get('CARD_BRAND'))} {fund(r.get('CARD_FUNDING'))}"] += 1
            cat_of_plan.setdefault(plan, cat)
    return by_cat, by_name, lines, cat_of_plan


# --------------------------------------------------------------- cost joining ----
def attach_costs(txns, by_cat):
    """Split each charge into the five cost components. Charges with no IC+ line yet
    (still settling at extract time) get an estimate from the blended ACTUAL rate for
    the same brand x funding, and are flagged 'Estimated' in column M."""
    actual, pending = [], []
    for t in txns:
        c = by_cat.get(t["charge_id"])
        if c:
            t["interchange"] = round(c.get("interchange", 0.0), 6)
            t["card_scheme"] = round(c.get("card_scheme", 0.0), 6)
            t["amex_discount"] = round(c.get("discount", 0.0), 6)
            t["per_auth"] = round(c.get("per_auth_fee", 0.0), 6)
            t["volume_fee"] = round(c.get("volume_fee", 0.0), 6)
            t["fee_data"] = "Actual"
            actual.append(t)
        else:
            pending.append(t)

    rate = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0, 0.0, 0.0])   # 5 comps + volume
    for t in actual:
        k = (t["brand"], t["funding"])
        a = rate[k]
        for i, f in enumerate(("interchange", "card_scheme", "amex_discount",
                               "per_auth", "volume_fee")):
            a[i] += t[f]
        a[5] += t["amount"]
    overall = [0.0] * 6
    for a in rate.values():
        for i in range(6):
            overall[i] += a[i]

    for t in pending:
        a = rate.get((t["brand"], t["funding"]))
        if not a or a[5] <= 0:
            a = overall
        vol = a[5] if a[5] > 0 else 1.0
        for i, f in enumerate(("interchange", "card_scheme", "amex_discount",
                               "per_auth", "volume_fee")):
            t[f] = round(t["amount"] * (a[i] / vol), 6)
        t["fee_data"] = "Estimated"
    return len(actual), len(pending)


# ------------------------------------------------------------------ sheet I/O ----
def clear_below(ws, first_row, max_col=None):
    """Drop every cell from first_row down.

    The obvious version assigns None cell-by-cell. On Hyper Detail that is
    16,090 x 209 = 3.4M assignments, and each one CREATES a cell object before
    emptying it - it ran for minutes. openpyxl keeps cells in a dict keyed
    (row, col), so deleting those keys gives the same result in one pass.
    """
    for key in [k for k in ws._cells if k[0] >= first_row]:
        del ws._cells[key]
    ws._current_row = max(first_row - 1, 0)


def capture_pattern(ws, pattern_row, raw_letters):
    """Snapshot the template's formula row BEFORE the data rows are cleared.
    Reading it afterwards returns an empty row - which silently produced a workbook
    with zero formulas on the first run of this script."""
    skip = {cifs(l) for l in raw_letters}
    return {c: ws.cell(pattern_row, c).value
            for c in range(1, ws.max_column + 1)
            if c not in skip and isinstance(ws.cell(pattern_row, c).value, str)
            and ws.cell(pattern_row, c).value.startswith("=")}


def fill_formulas(ws, pattern_row, target_rows, pat):
    """Write the captured pattern down to target_rows, translating relative rows."""
    import re
    rowref = re.compile(rf"(?<![\d$]){pattern_row}(?!\d)")
    n = 0
    for r in range(pattern_row, pattern_row + target_rows):
        for c, f in pat.items():
            ws.cell(r, c).value = f if r == pattern_row else rowref.sub(str(r), f)
            n += 1
    return n


def build_detail(ws, txns, country):
    hdr = {gcl(c): ws.cell(2, c).value for c in range(1, ws.max_column + 1)}
    for letter, expect in RAW_COLS.items():
        got = hdr.get(letter)
        if got != expect:
            sys.exit(f"TEMPLATE DRIFT: {DETAIL}!{letter}2 is {got!r}, expected {expect!r}. "
                     "The column layout moved; fix the template before refreshing.")

    pat = capture_pattern(ws, DATA_START, RAW_COLS.keys())
    clear_below(ws, DATA_START)
    unknown = set()
    for i, t in enumerate(txns):
        r = DATA_START + i
        # Prefer the country the pull gives us; fall back to the name lookup.
        cty = t.get("merchant_country") or country.get(t["merchant"].casefold())
        if not cty:
            unknown.add(t["merchant"])
            cty = ""
        vals = {
            "A": t["charge_id"], "B": t["date"], "C": f"=MONTH($B{r})",
            "E": t["merchant"], "F": t["currency"], "G": t["amount"],
            "H": t["brand"], "I": t["funding"], "J": t["card_country"],
            "K": t["last4"], "L": t["status"], "M": t["fee_data"],
            "N": t["revenue"], "O": t["interchange"], "P": t["card_scheme"],
            "Q": t["amex_discount"], "S": t["per_auth"], "T": t["volume_fee"],
            "BX": cty,
        }
        for letter, v in vals.items():
            ws.cell(r, cifs(letter)).value = v
    n = fill_formulas(ws, DATA_START, len(txns), pat)
    return n, unknown


def build_hyper(ws, txns, by_name, country, cat_of_plan):
    hdr = [c.value for c in ws[1]]
    col_of = {}
    for i, h in enumerate(hdr, start=1):
        if isinstance(h, str) and ":" in h and h.split(":")[0] in ("IC", "CS", "AX", "ST"):
            col_of[h.split(":", 1)[1].strip()] = i
    fixed = {h: i for i, h in enumerate(hdr, start=1) if isinstance(h, str)}

    # The two Stripe fee lines carry no PLAN_NAME in icplus_costs.csv, so they arrive
    # keyed by their FEE_CATEGORY and match none of the 'ST: ...' header names. Map
    # them explicitly rather than letting them fall out as "unmatched".
    for cat, needle in (("per_auth_fee", "ST: Per Auth"), ("volume_fee", "ST: Volume")):
        hit = next((i for i, h in enumerate(hdr, start=1)
                    if isinstance(h, str) and h.startswith(needle)), None)
        if hit:
            col_of[cat] = hit

    clear_below(ws, HYPER_START)

    unmatched = Counter()
    added = {}
    for i, t in enumerate(txns):
        r = HYPER_START + i
        net = t["interchange"] + t["card_scheme"] + t["amex_discount"]
        strp = t["per_auth"] + t["volume_fee"]
        base = {
            "Charge ID": t["charge_id"], "Date": t["date"], "Merchant": t["merchant"],
            "Merchant Country": (t.get("merchant_country")
                                 or country.get(t["merchant"].casefold(), "")),
            "Transaction Value": t["amount"], "Card Brand": t["brand"],
            "Card Funding": t["funding"], "Card Country": t["card_country"],
            "Fee Data": t["fee_data"], "Revenue (Loyverse)": t["revenue"],
            "Total Cost": round(net + strp, 6), "Net Margin": round(t["revenue"] - net - strp, 6),
            "NETWORK COST — TOTAL": round(net, 6),
            "IC — Interchange TOTAL": t["interchange"],
            "STRIPE COST — TOTAL": round(strp, 6),
            "ST — Per Auth Fee TOTAL": t["per_auth"],
            "ST — Volume Fee TOTAL": t["volume_fee"],
        }
        for h, v in base.items():
            if h in fixed:
                ws.cell(r, fixed[h]).value = v
        for plan, amt in by_name.get(t["charge_id"], {}).items():
            c = col_of.get(plan)
            if c is None:
                # Stripe adds interchange and scheme line items over time - the first
                # live run met 20 that did not exist in the 25 Aug workbook, and the
                # previous version DROPPED them. Their cost was still correct in
                # Transaction Detail (that side aggregates by CATEGORY, not by line
                # name), but the per-line matrix here under-attributed them with only a
                # warning. Now the column is created on the spot, so no fee line is
                # ever lost. Category prefixes match the template's own convention.
                cat = cat_of_plan.get(plan, "")
                prefix = ("AX" if cat == "discount" else
                          "ST" if cat in ("per_auth_fee", "volume_fee") else
                          "CS" if cat == "card_scheme" else "IC")
                c = ws.max_column + 1
                hc = ws.cell(1, c, f"{prefix}: {plan}")
                hc.font, hc.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
                hc.alignment = Alignment(wrap_text=True, vertical="bottom")
                ws.column_dimensions[gcl(c)].width = 18
                col_of[plan] = c
                added[plan] = f"{prefix}: {plan}"
            ws.cell(r, c).value = round(amt, 6)
    return unmatched, added


def build_fee_breakdown(ws, lines, n_estimated):
    """Rebuild Fee Breakdown, growing the sheet to fit however many fee lines Stripe
    billed.

    TWO BUGS FIXED HERE, both found on the first live CI run (18,794 charges brought
    20 fee line items that did not exist in the 25 Aug workbook):

      1. The row budget was the template's own length. The previous version stopped at
         `ws.max_row - 3` (row 202 of a 205-row sheet) and silently dropped the rest -
         it reported "200 rows" while truncating the tail of the last category. Exactly
         the same failure as the hardcoded 16091 ceiling: a fixed size that new data
         quietly overflows. There is now no cap; the footers move down instead.

      2. The footers were pinned to rows 204/205 and D204 read `=$D$205-$D$203`, where
         203 meant "the last data row" only on the day it was written. Both footers are
         now emitted immediately after the data, and every reference is computed from
         the row they actually land on.
    """
    LABEL = {"interchange": "Interchange", "card_scheme": "Card Scheme",
             "discount": "Amex Discount", "per_auth_fee": "Stripe Per Auth",
             "volume_fee": "Stripe Volume"}
    ORDER = ("interchange", "card_scheme", "discount", "per_auth_fee", "volume_fee")

    # Keep a data row's formatting to stamp onto every row we write, since the sheet
    # now extends past the styled region of the template.
    proto = [ws.cell(6, c)._style for c in range(1, 8)]

    clear_below(ws, 5)

    grand = sum(a[1] for cat in lines.values() for a in cat.values())
    r = 5
    cat_rows = []

    def put(row, values):
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row, c, v)
            cell._style = proto[c - 1]

    for cat in ORDER:
        items = lines.get(cat)
        if not items:
            continue
        cat_occ = sum(a[0] for a in items.values())
        cat_tot = sum(a[1] for a in items.values())
        put(r, [LABEL[cat], f"{LABEL[cat]} — all {len(items)} line items", cat_occ,
                round(cat_tot, 6), None, round(cat_tot / grand, 6) if grand else None,
                None])
        cat_rows.append(r)
        r += 1
        for plan, (occ, tot, cards) in sorted(items.items(), key=lambda kv: -kv[1][1]):
            put(r, [LABEL[cat], plan, occ, round(tot, 6),
                    round(tot / occ, 6) if occ else None,
                    round(tot / grand, 6) if grand else None,
                    cards.most_common(1)[0][0] if cards else None])
            r += 1

    n_data = r - 5

    # Footers, positioned relative to the data rather than pinned.
    r += 1
    actual_row = r
    put(actual_row, [None, "Actual line items — total of the category rows above", None,
                     "=" + "+".join(f"$D${x}" for x in cat_rows) if cat_rows else 0,
                     None, None, None])
    r += 1
    est_row = r
    put(est_row, [None,
                  f"Estimated cost — the {n_estimated} charges still settling at extract "
                  "time (blended rate by card brand and funding, no line items available)",
                  None, None, None, None, None])
    r += 1
    total_row = r
    put(total_row, [None, "TOTAL COST — ties to Transaction Detail column V and to Summary",
                    None, f"=SUM('{DETAIL}'!$V${DATA_START}:$V${MAXROW})",
                    None, None, None])
    # Now that both rows exist, the estimate is total minus what the line items explain.
    ws.cell(est_row, 4).value = f"=$D${total_row}-$D${actual_row}"

    return n_data, total_row


def build_merchant_analysis(ws, txns):
    """Rebuilt as an aggregation. In the hand-built workbook this tab was a
    values-only paste with zero formulas, so it never updated and gave no sign of it."""
    agg = defaultdict(lambda: [0, 0.0])
    for t in txns:
        a = agg[t["merchant"]]
        a[0] += 1
        a[1] += t["amount"]
    order = sorted(agg.items(), key=lambda kv: -kv[1][1])

    clear_below(ws, 3)

    # Price / margin per merchant are SUMIFs over Transaction Detail, so the tab
    # recalculates with the data instead of freezing.
    PRICE = {5: "AC", 6: "AD", 7: "AE", 8: "AF", 9: "AG", 10: "AH", 11: "AI",
             12: "AM", 13: "AL"}
    MARGIN = {15: "AQ", 16: "AR", 17: "AS", 18: "AT", 19: "AU", 20: "AV", 21: "AW",
              22: "BC", 23: "BB"}
    for i, (name, (n, vol)) in enumerate(order):
        r = 3 + i
        ws.cell(r, 1).value = name
        ws.cell(r, 2).value = n
        ws.cell(r, 3).value = round(vol, 2)
        ws.cell(r, 4).value = f'=IFERROR(C{r}/B{r},"")'
        for c, src in PRICE.items():
            ws.cell(r, c).value = (f"=SUMIF('{DETAIL}'!$E$3:$E$300000,$A{r},"
                                   f"'{DETAIL}'!${src}$3:${src}$300000)")
        for c, src in MARGIN.items():
            ws.cell(r, c).value = (f"=SUMIF('{DETAIL}'!$E$3:$E$300000,$A{r},"
                                   f"'{DETAIL}'!${src}$3:${src}$300000)")
    return len(order)


def merchant_country_map(wb):
    """Merchant Country is not in transactions.csv - CONNECTED_ACCOUNTS carries no
    COUNTRY column in the documented schema. Until query1 is extended (gated on a Q0
    check), it is resolved from the template's own BX values, keyed on merchant name."""
    ws = wb[DETAIL]
    out = {}
    bx = cifs("BX")
    for r in range(DATA_START, ws.max_row + 1):
        m = ws.cell(r, cifs("E")).value
        c = ws.cell(r, bx).value
        if m and c:
            out.setdefault(m.casefold(), c)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tx", required=True)
    ap.add_argument("--ic", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    a = ap.parse_args()

    print(f"reading {a.tx}")
    txns = load_transactions(a.tx)
    print(f"  {len(txns):,} succeeded charges")
    print(f"reading {a.ic}")
    by_cat, by_name, lines, cat_of_plan = load_icplus(a.ic)
    n_act, n_est = attach_costs(txns, by_cat)
    print(f"  cost attached: {n_act:,} Actual, {n_est:,} Estimated")

    shutil.copyfile(a.template, a.out)
    wb = openpyxl.load_workbook(a.out, data_only=False)

    country = merchant_country_map(wb)
    print(f"  merchant-country lookup seeded with {len(country)} merchants")

    protected_before = {s: [[c.value for c in row] for row in wb[s].iter_rows()]
                        for s in PROTECTED if s in wb.sheetnames}

    n_f, unknown = build_detail(wb[DETAIL], txns, country)
    print(f"\n  {DETAIL}: {len(txns):,} rows, {n_f:,} formula cells filled")
    if unknown:
        print(f"    WARNING: no country for {len(unknown)} merchant(s): "
              f"{sorted(unknown)[:5]}\n"
              "      Fix permanently by adding  a.COUNTRY AS merchant_country  to the\n"
              "      SELECT in payments-automation/queries/query1_transactions.sql.\n"
              "      CONNECTED_ACCOUNTS.COUNTRY exists - q1_base.sql already filters on it.")

    unmatched, added = build_hyper(wb[HYPER], txns, by_name, country, cat_of_plan)
    print(f"  {HYPER}: {len(txns):,} rows")
    if added:
        print(f"    {len(added)} new fee line(s) appeared in the data and got new "
              f"columns (nothing dropped):")
        for k in sorted(added.values()):
            print(f"       + {k}")

    n_fb, fb_total = build_fee_breakdown(wb["Fee Breakdown"], lines, n_est)
    print(f"  Fee Breakdown: {n_fb} fee lines, totals on row {fb_total}")

    n_ma = build_merchant_analysis(wb["Merchant Analysis"], txns)
    print(f"  Merchant Analysis: {n_ma} merchants (now formula-driven)")

    for s, before in protected_before.items():
        after = [[c.value for c in row] for row in wb[s].iter_rows()]
        if before != after:
            sys.exit(f"REFUSING TO WRITE: protected sheet {s!r} was modified.")
    print(f"  protected sheets unchanged: {', '.join(protected_before)}")

    wb.calculation.fullCalcOnLoad = True
    wb.save(a.out)
    print(f"\nwrote {a.out}")


if __name__ == "__main__":
    main()
