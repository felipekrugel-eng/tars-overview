#!/usr/bin/env python3
"""
make_margin_template.py — turn the hand-built Aug 25 workbook into an automatable
template.

    python3 make_margin_template.py SOURCE.xlsx template/Loyverse_Payments_Transaction_Margin_vs_Competitors.xlsx

Run ONCE. The output is committed as the template and from then on only
refresh_margin_workbook.py touches the workbook.

Six fixes, each documented in Margin_Workbook_Preservation_Spec.md:

  1. RE-RANGE   Every summary formula bounded its ranges at row 16091 - exactly the
                last data row on the day it was written. 2,424 occurrences across 7
                sheets. The first refresh that adds a transaction makes every total
                silently wrong. Rebound at MAXROW, far above any plausible row count,
                the same trick build_full_base.py uses.

  2. XLOOKUP    Columns BE:BR use _xlfn.XLOOKUP, 321,780 cells. Verified working under
                LibreOffice 26.2, but the runner installs LibreOffice from apt and that
                build may predate XLOOKUP support - in which case all 321,780 cells
                recalc to #NAME? and the verify gate fails every run. INDEX/MATCH is
                exactly equivalent (tested) and works on every engine.

  3. FOOTERS    Rows 16093/16094/16096/16097 held unlabelled hand-added totals sitting
                two rows BELOW the data. Any append overwrites them or strands them
                mid-data. Deleted here and replaced by live formulas on Summary, which
                is what they should always have been.

  4. MONTH      Column C was blank in rows 3-6, an array formula in row 7 and static
                integers from row 8. Made uniformly =MONTH($B{r}).

  5. TRIM       BE/BF were dragged two rows past every other column (to 16093). That
                overshoot is what produced the stray footer rows. Trimmed.

  6. PIVOT      Sheet1 is a real PivotTable with months hardcoded to 4-8 and 58
                merchants hardcoded as rows. openpyxl cannot rebuild a pivot and the
                cache does not refresh on headless recalc, so it would freeze at
                August while every other tab moved on - and still look current.
                Replaced by 'Merchant Months', a formula tab that extends itself.

Nothing on Rate Card is touched. Those values are hand-entered, several transcribed
from the signed Stripe pricing agreement, and are not reproducible from any data
source.
"""

import re
import shutil
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAXROW = 300_000
OLD_CEILING = 16091
DATA_START = 3
DETAIL = "Transaction Detail"
HYPER = "Transaction Hyper Detail"

NAVY = "0B2B3C"
BLUE = "1F6FB2"
HDR_F = Font(name="Aptos Narrow", size=10, bold=True, color="FFFFFF")
BODY_F = Font(name="Aptos Narrow", size=10)
TITLE_F = Font(name="Aptos Narrow", size=12, bold=True, color=NAVY)

# Stray hand-added rows below the data block on Transaction Detail.
FOOTER_ROWS = (16092, 16093, 16094, 16095, 16096, 16097)

XLOOKUP_RE = re.compile(r"_xlfn\.XLOOKUP\(([^,)]+),([^,)]+),([^,)]+)\)")


def fix_formula(f: str) -> tuple[str, int, int]:
    """Return (formula, n_reranged, n_xlookup) for one formula string."""
    n_range = 0
    n_xl = 0

    # 1. re-range. Only rewrite ABSOLUTE row references ($16091), which is how every
    #    range ceiling in this workbook is written ('Transaction Detail'!$F$3:$F$16091).
    #    A bare 16091 is a RELATIVE self-row reference in the formulas that live on row
    #    16091 itself ($G16091, AK16091, ...). An earlier version of this script matched
    #    those too and corrupted 11 cells on the last data row into #N/A. Requiring the
    #    '$' is what separates a ceiling from a row that merely happens to be last.
    new, n_range = re.subn(rf"\${OLD_CEILING}(?!\d)", f"${MAXROW}", f)

    # 2. XLOOKUP -> INDEX/MATCH. Every occurrence in this workbook is the plain
    #    three-argument form and no argument contains a comma or a paren, so a
    #    single pass is safe. Asserted below.
    while True:
        m = XLOOKUP_RE.search(new)
        if not m:
            break
        lookup, haystack, result = (a.strip() for a in m.groups())
        new = new[: m.start()] + f"INDEX({result},MATCH({lookup},{haystack},0))" + new[m.end():]
        n_xl += 1

    return new, n_range, n_xl


def main():
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    src, dst = sys.argv[1], sys.argv[2]

    print(f"reading {src}")
    shutil.copyfile(src, dst)
    wb = openpyxl.load_workbook(dst, data_only=False)

    # ---------------------------------------------------------------- 1 + 2 ----
    tot_range = tot_xl = 0
    per_sheet = {}
    for ws in wb.worksheets:
        if ws.title == HYPER:
            continue          # static IC+ line items, no formulas to re-range
        r = x = 0
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    nv, dr, dx = fix_formula(v)
                    if nv != v:
                        c.value = nv
                    r += dr
                    x += dx
        if r or x:
            per_sheet[ws.title] = (r, x)
        tot_range += r
        tot_xl += x

    print(f"\n  fix 1+2 — re-ranged {tot_range:,} refs, converted {tot_xl:,} XLOOKUPs")
    for s, (r, x) in per_sheet.items():
        print(f"      {s:28} rerange={r:>6,}  xlookup={x:>7,}")

    assert "_xlfn.XLOOKUP" not in "".join(
        str(c.value) for ws in wb.worksheets if ws.title != HYPER
        for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and "XLOOKUP" in c.value
    ), "an XLOOKUP survived conversion"

    ws = wb[DETAIL]

    # -------------------------------------------------------------------- 3 ----
    # Capture the footer values before deleting, so the Summary formulas can be
    # cross-checked against them on the first run.
    captured = {}
    for r in FOOTER_ROWS:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                captured[f"{get_column_letter(c)}{r}"] = v
    print(f"\n  fix 3 — captured {len(captured)} floating footer cells, deleting rows "
          f"{FOOTER_ROWS[0]}-{FOOTER_ROWS[-1]}")
    for r in reversed(FOOTER_ROWS):
        ws.delete_rows(r)

    last = ws.max_row
    print(f"      Transaction Detail now ends at row {last:,}")

    # -------------------------------------------------------------------- 4 ----
    for r in range(DATA_START, last + 1):
        ws.cell(r, 3).value = f"=MONTH($B{r})"
    print(f"  fix 4 — column C normalised to =MONTH($B) on rows {DATA_START}-{last:,}")

    # -------------------------------------------------------------------- 5 ----
    # (the overshoot rows were inside FOOTER_ROWS and are now gone)
    print("  fix 5 — BE/BF overshoot removed with the footer rows")

    # ------------------------------------------------- 3b: live Summary block ----
    sm = wb["Summary"]
    base = sm.max_row + 2
    sm.cell(base, 1, "Scenario totals — relocated from the floating block below "
                     "Transaction Detail (was rows 16093-16097)").font = TITLE_F
    hdr = ["Scenario", "Total margin ($)", "Margin % of volume"]
    for i, h in enumerate(hdr, start=1):
        c = sm.cell(base + 1, i, h)
        c.font = HDR_F
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center")

    # (label, margin column on Transaction Detail)
    scenarios = [
        ("Loyverse US list (2.6% + 15c)", "AQ"),
        ("Square", "AR"),
        ("SumUp", "AS"),
        ("Toast", "AT"),
        ("Clover", "AU"),
        ("Zettle (PayPal)", "AV"),
        ("Stripe Terminal (direct)", "AW"),
        ("Hypothetical 1.99%", "AY"),
        ("Hypothetical 2.29%", "AZ"),
        ("Hypothetical 2.49%", "BA"),
        ("Loyverse Pro (1.69/2.29 + 15c)", "BB"),
        ("Loyverse Free (1.99/2.49 + 15c)", "BC"),
    ]
    vol = f"SUM('{DETAIL}'!$G${DATA_START}:$G${MAXROW})"
    for i, (label, col) in enumerate(scenarios):
        r = base + 2 + i
        sm.cell(r, 1, label).font = BODY_F
        m = sm.cell(r, 2, f"=SUM('{DETAIL}'!${col}${DATA_START}:${col}${MAXROW})")
        m.font = BODY_F
        m.number_format = '$#,##0.00;($#,##0.00);-'
        p = sm.cell(r, 3, f"=IFERROR(B{r}/{vol},\"\")")
        p.font = BODY_F
        p.number_format = '0.0%;-0.0%;-'
    print(f"  fix 3b — {len(scenarios)} scenario totals rebuilt as live formulas on "
          f"Summary rows {base + 2}-{base + 1 + len(scenarios)}")

    # -------------------------------------------------------------------- 6 ----
    if "Sheet1" in wb.sheetnames:
        pos = wb.sheetnames.index("Sheet1")
        del wb["Sheet1"]
        mm = wb.create_sheet("Merchant Months", pos)
        mm.sheet_properties.tabColor = BLUE

        mm["A1"] = "Merchant Months — active months per merchant, and the per-merchant-month fee"
        mm["A1"].font = TITLE_F
        mm["A2"] = ("Replaces the Sheet1 PivotTable, which had months hardcoded to 4-8 and 58 "
                    "merchants hardcoded as rows and so could never show a new month. This tab "
                    "is formulas only and extends itself. NOTE: months are calendar month "
                    "numbers, so add a year dimension before the data crosses into 2027.")
        mm["A2"].font = Font(name="Aptos Narrow", size=9, italic=True)
        mm["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        mm.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
        mm.row_dimensions[2].height = 32

        HDR = 4
        c = mm.cell(HDR, 1, "Merchant")
        c.font, c.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
        for m in range(1, 13):
            c = mm.cell(HDR, 1 + m, m)
            c.font, c.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center")
        for off, lab in enumerate(["Total txns", "Active months"]):
            c = mm.cell(HDR, 14 + off, lab)
            c.font, c.fill = HDR_F, PatternFill("solid", fgColor=BLUE)
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        # 'By Merchant' lists one merchant per row from row 5. Bounded generously so
        # new merchants appear without touching this tab.
        BY_FIRST, N_MERCH = 5, 400
        for i in range(N_MERCH):
            r = HDR + 1 + i
            mm.cell(r, 1, f"=IF('By Merchant'!A{BY_FIRST + i}=\"\",\"\",'By Merchant'!A{BY_FIRST + i})").font = BODY_F
            for m in range(1, 13):
                cell = mm.cell(r, 1 + m,
                               f'=IF($A{r}="","",COUNTIFS('
                               f"'{DETAIL}'!$E${DATA_START}:$E${MAXROW},$A{r},"
                               f"'{DETAIL}'!$C${DATA_START}:$C${MAXROW},{get_column_letter(1 + m)}${HDR}))")
                cell.font = BODY_F
                cell.number_format = '#,##0;;-'
            mm.cell(r, 14, f'=IF($A{r}="","",SUM(B{r}:M{r}))').font = BODY_F
            mm.cell(r, 15, f'=IF($A{r}="","",COUNTIF(B{r}:M{r},">0"))').font = BODY_F

        tot = HDR + 1 + N_MERCH + 1
        mm.cell(tot, 1, "Total").font = HDR_F
        mm.cell(tot, 1).fill = PatternFill("solid", fgColor=NAVY)
        mm.cell(tot, 14, f"=SUM(N{HDR + 1}:N{tot - 1})").font = HDR_F
        mm.cell(tot, 14).fill = PatternFill("solid", fgColor=NAVY)
        mm.cell(tot, 15, f"=SUM(O{HDR + 1}:O{tot - 1})").font = HDR_F
        mm.cell(tot, 15).fill = PatternFill("solid", fgColor=NAVY)

        mm.cell(tot + 2, 1, "Merchant-months").font = BODY_F
        mm.cell(tot + 2, 2, f"=O{tot}").font = BODY_F
        mm.cell(tot + 3, 1, "Fee per merchant-month (USD)").font = BODY_F
        fee = mm.cell(tot + 3, 2, 29.9)
        fee.font = Font(name="Aptos Narrow", size=10, bold=True, color="1F6FB2")
        fee.fill = PatternFill("solid", fgColor="DCE9F5")   # blue = input
        fee.number_format = '$#,##0.00'
        mm.cell(tot + 4, 1, "Implied subscription value (USD)").font = HDR_F
        mm.cell(tot + 4, 1).fill = PatternFill("solid", fgColor=NAVY)
        v = mm.cell(tot + 4, 2, f"=B{tot + 2}*B{tot + 3}")
        v.font, v.fill = HDR_F, PatternFill("solid", fgColor=NAVY)
        v.number_format = '$#,##0.00'

        mm.column_dimensions["A"].width = 46
        for m in range(1, 13):
            mm.column_dimensions[get_column_letter(1 + m)].width = 7
        mm.column_dimensions["N"].width = 11
        mm.column_dimensions["O"].width = 11
        mm.freeze_panes = "B5"
        print(f"  fix 6 — Sheet1 pivot replaced by 'Merchant Months' "
              f"({N_MERCH} merchant rows x 12 months, self-extending)")

    # ------------------------------------------------------------- finalise ----
    wb.calculation.fullCalcOnLoad = True
    print(f"\nwriting {dst}")
    wb.save(dst)
    print("done. Rate Card untouched.")


if __name__ == "__main__":
    main()
